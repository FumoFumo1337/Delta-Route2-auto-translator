"""Translate complete workbook dialog windows through the DeepL API.

Every translated dialog window is appended to a JSONL cache keyed by its D/C
identifier, translation method, and source signature, so repeated Japanese in
unrelated scenes cannot overwrite its context and an interrupted run resumes
without paying twice. Legacy source-keyed entries are migrated without another
API call. Each dialog window is sent as one continuous text. The result is then
wrapped back to the original line count at word boundaries, following the
relative lengths of the Japanese source lines; a window that comes back without
enough spaces to wrap at is cut on punctuation instead, so that no line of it
is left empty. If DeepL visibly loops while translating a joined multi-line
window, that window alone is retried with one API text per source line.
"""

from __future__ import annotations

import argparse
import html
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
import xml.etree.ElementTree as ET
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install openpyxl"
    ) from exc


DEFAULT_API_URL = os.environ.get(
    "DEEPL_API_URL", "https://api-free.deepl.com/v2/translate"
)
DEEPL_MAX_TEXTS = 50
# Content retries are separate from request_deepl's transport retries. A broken
# answer can be HTTP-successful every time, so it needs its own hard quota guard.
DEFAULT_LOOP_RETRIES = 0
UNRESOLVED_LOOPS_FILE_TEMPLATE = "deepl_unresolved_loops.{language}.json"


def unresolved_loop_report_name(target_lang: str) -> str:
    return UNRESOLVED_LOOPS_FILE_TEMPLATE.format(language=target_lang.strip().lower())


# How a translation was produced, recorded per cache entry. A line of a message
# window is one slice of a sentence translated whole; a menu option is a text of
# its own. The same Japanese can appear as both - two lines do in the Reika
# script - and the two results are not interchangeable, so they are cached apart
# instead of one silently satisfying the lookup for the other.
CACHE_METHOD = "dialog-proportional-v1"
CHOICE_CACHE_METHOD = "choice-line-v1"
# Current scenario cache entries describe the same unit sent to DeepL: a whole
# D/C window. The old per-source markers above remain readable for migration and
# for delta_menu.py's independent catalog-caption cache.
DIALOG_CACHE_METHOD = "dialog-window-v2"
CHOICE_DIALOG_CACHE_METHOD = "choice-window-v2"

PROTECTED_TOKEN_RE = re.compile(
    r"\[[^\]\r\n]*\]"
    r"|\$\{[^}\r\n]+\}"
    r"|%\([^)\r\n]+\)[#0\- +]?\d*(?:\.\d+)?[A-Za-z]"
    r"|%[#0\- +]?\d*(?:\.\d+)?[A-Za-z%]"
    r"|\\[nrt\\]"
)
# Where a translation without spaces may still be wrapped. Each atom carries
# the punctuation it ends on, so an atom list joins back into the text it came
# from and a line break lands after a pause rather than inside a word.
DIALOG_ATOM_SEPARATORS = r"\s…‥。、，,.!?！？—―・:;-"
DIALOG_ATOM_RE = re.compile(
    PROTECTED_TOKEN_RE.pattern
    + f"|[^{DIALOG_ATOM_SEPARATORS}]+[{DIALOG_ATOM_SEPARATORS}]*"
    + f"|[{DIALOG_ATOM_SEPARATORS}]+"
)
DEEPL_MARKER_RE = re.compile(
    r'<x\s+id=["\'](\d+)["\']\s*(?:/\s*>|>\s*</x\s*>)', re.IGNORECASE
)
XML_INVALID_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\uFFFE\uFFFF]")
# D is a message window, C a choice menu. Both group lines the same way, so the
# translator treats them alike; the letter is there for the reviewer.
DIALOG_REF_RE = re.compile(r"\b([CD]\d{5})\s+(\d+)/(\d+)\b")
LOOP_WORD_RE = re.compile(r"[^\W\d_]+(?:['’][^\W\d_]+)?", re.UNICODE)


@dataclass(frozen=True)
class WorkbookRow:
    source: str
    worksheet: Any
    row_index: int
    target_column: int
    dialog_line: str = ""


@dataclass
class DialogJob:
    key: str
    rows: list[WorkbookRow]
    wanted_sources: set[str]
    # Whether the lines go as texts of their own rather than as one wrapped
    # message. The engine wraps a message across up to three source lines, so a
    # window has to be translated whole and wrapped back or the halves stop
    # agreeing. A menu is the opposite: its lines are alternatives the reader
    # chooses between, and joining them produces one sentence sliced across the
    # options - `Завтра проверим зоны. Завтра` against `займёмся вибратором.`
    # standalone_windows decides this; a menu is never the other thing.
    per_line: bool = False

    def __post_init__(self) -> None:
        if self.key.startswith("C") and not self.per_line:
            raise ValueError(
                f"Menu {self.key} cannot be joined into one text: an option is "
                "only ever translated on its own"
            )

    @property
    def text_count(self) -> int:
        """How many texts this window costs inside one request."""
        return len(self.rows) if self.per_line else 1

    @property
    def character_count(self) -> int:
        sources = [row.source for row in self.rows]
        if self.per_line:
            return sum(len(source.strip()) for source in sources)
        return len(join_dialog_source(sources, "JA"))


def row_cache_method(source: str, standalone: set[str]) -> str:
    """Which cache series a row belongs to.

    `standalone` is what standalone_sources returns. A line shared between a
    menu and a message is one of them: the self-contained wording reads
    correctly as a message line too, while a fragment taken from the middle of a
    sentence does not read as an option at all.
    """
    return CHOICE_CACHE_METHOD if source in standalone else CACHE_METHOD


def cell_text(value: Any) -> str:
    return "" if value is None else str(value)


def find_header_cell(ws: Any, header: str) -> int:
    wanted = header.casefold()
    for col in range(1, ws.max_column + 1):
        if cell_text(ws.cell(1, col).value).casefold() == wanted:
            return col
    raise ValueError(f"Sheet {ws.title!r} does not have header {header!r}")


def load_cache(path: Path, target_lang: str) -> dict[str, dict[str, str]]:
    """Every usable entry, as source -> method -> translation.

    Keeping the method rather than filtering on one is what let the menus be
    re-translated without touching the rest: their sources have an entry under
    the old method and none under the new one, so only they come out uncached.
    """
    cache: dict[str, dict[str, str]] = {}
    wanted_language = target_lang.upper()
    known = {CACHE_METHOD, CHOICE_CACHE_METHOD}
    if not path.exists():
        return cache
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                print(f"Skipping a damaged cache line in {path.name}", file=sys.stderr)
                continue
            item_language = str(item.get("target_lang", "RU")).upper()
            method = str(item.get("method", ""))
            if item_language == wanted_language and method in known:
                cache.setdefault(str(item["source"]), {})[method] = str(
                    item["translation"]
                )
    return cache


def cache_for_rows(
    stored: dict[str, dict[str, str]],
    rows: list[WorkbookRow],
    standalone: set[str] | None = None,
) -> dict[str, str]:
    """Collapses the stored entries to the one translation each row may reuse."""
    if standalone is None:
        standalone = standalone_sources(collect_windows(rows, require_complete=False))
    methods = {row.source: row_cache_method(row.source, standalone) for row in rows}
    return {
        source: entries[methods[source]]
        for source, entries in stored.items()
        if methods.get(source) in entries
    }


def append_cache(
    path: Path, source: str, translation: str, target_lang: str, method: str
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(
            json.dumps(
                {
                    "source": source,
                    "translation": translation,
                    "target_lang": target_lang.upper(),
                    "method": method,
                },
                ensure_ascii=False,
                separators=(",", ":"),
            )
            + "\n"
        )


DialogCacheKey = tuple[str, str, tuple[str, ...]]


def dialog_cache_method(job: DialogJob) -> str:
    return CHOICE_DIALOG_CACHE_METHOD if job.per_line else DIALOG_CACHE_METHOD


def load_dialog_cache(
    path: Path, target_lang: str
) -> dict[DialogCacheKey, tuple[str, ...]]:
    """Loads contextual D/C-window entries, guarded by their source signature."""
    cache: dict[DialogCacheKey, tuple[str, ...]] = {}
    wanted_language = target_lang.upper()
    known = {DIALOG_CACHE_METHOD, CHOICE_DIALOG_CACHE_METHOD}
    if not path.exists():
        return cache
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                # load_cache reports the same damaged physical line when both
                # formats are read by main(), so avoid printing it twice here.
                continue
            method = str(item.get("method", ""))
            language = str(item.get("target_lang", "RU")).upper()
            dialog = str(item.get("dialog", ""))
            sources = item.get("source_lines")
            translations = item.get("translation_lines")
            if language != wanted_language or method not in known:
                continue
            if not re.fullmatch(r"[CD]\d{5}", dialog):
                continue
            if not isinstance(sources, list) or not isinstance(translations, list):
                continue
            if not sources or len(sources) != len(translations):
                continue
            source_signature = tuple(str(source) for source in sources)
            translated_lines = tuple(str(value) for value in translations)
            cache[(dialog, method, source_signature)] = translated_lines
    return cache


def append_dialog_cache(
    path: Path,
    job: DialogJob,
    translations: list[str] | tuple[str, ...],
    target_lang: str,
) -> None:
    """Appends one API translation unit rather than context-free line entries."""
    if len(translations) != len(job.rows):
        raise ValueError(
            f"Cache entry for {job.key} has {len(translations)} translations "
            f"for {len(job.rows)} source lines"
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(
            json.dumps(
                {
                    "dialog": job.key,
                    "source_lines": [row.source for row in job.rows],
                    "translation_lines": list(translations),
                    "target_lang": target_lang.upper(),
                    "method": dialog_cache_method(job),
                },
                ensure_ascii=False,
                separators=(",", ":"),
            )
            + "\n"
        )


def cache_for_dialog_jobs(
    stored: dict[DialogCacheKey, tuple[str, ...]],
    legacy: dict[str, dict[str, str]],
    jobs: list[DialogJob],
) -> tuple[dict[str, tuple[str, ...]], set[str]]:
    """Resolves exact contextual entries and safely migrates old source keys.

    The source signature is part of the lookup even though the D/C identifier
    is the primary identity. If extraction renumbers the story, an old D number
    can never silently translate a different window.
    """
    resolved: dict[str, tuple[str, ...]] = {}
    migrated: set[str] = set()
    for job in jobs:
        sources = tuple(row.source for row in job.rows)
        contextual = stored.get((job.key, dialog_cache_method(job), sources))
        if contextual is not None:
            resolved[job.key] = contextual
            continue

        legacy_lines: list[str] = []
        for source in sources:
            entries = legacy.get(source, {})
            if job.key.startswith("C"):
                # Never revive the joined-message slice that originally made a
                # menu option read as half of somebody else's sentence.
                translation = entries.get(CHOICE_CACHE_METHOD)
            elif job.per_line:
                # A D window can become per-line only because it shares a source
                # with a choice. Prefer the corrected option wording for shared
                # sources and retain the old dialog translation for its other
                # lines; this migrates the complete paid cache without another
                # request while preserving the ownership rule.
                translation = entries.get(
                    CHOICE_CACHE_METHOD, entries.get(CACHE_METHOD)
                )
            else:
                translation = entries.get(CACHE_METHOD)
            if translation is None:
                break
            legacy_lines.append(translation)
        if len(legacy_lines) == len(sources):
            resolved[job.key] = tuple(legacy_lines)
            migrated.add(job.key)
    return resolved, migrated


def iter_batches(
    jobs: list[DialogJob], max_chars: int, max_items: int
) -> list[list[DialogJob]]:
    batches: list[list[DialogJob]] = []
    current: list[DialogJob] = []
    current_chars = 0
    # A menu costs one text per option rather than one per window, so the item
    # limit counts texts. For windows, which are all one text, this is the same
    # arithmetic the limit has always done.
    current_texts = 0

    for item in jobs:
        item_chars = item.character_count
        item_texts = item.text_count
        if current and (
            current_chars + item_chars > max_chars
            or current_texts + item_texts > max_items
        ):
            batches.append(current)
            current = []
            current_chars = 0
            current_texts = 0
        current.append(item)
        current_chars += item_chars
        current_texts += item_texts

    if current:
        batches.append(current)
    return batches


def protect_markup(text: str) -> tuple[str, list[str]]:
    text = XML_INVALID_CHAR_RE.sub("\uFFFD", text)
    parts: list[str] = []
    tokens: list[str] = []
    offset = 0

    for match in PROTECTED_TOKEN_RE.finditer(text):
        parts.append(html.escape(text[offset : match.start()]))
        marker_id = len(tokens)
        tokens.append(match.group(0))
        parts.append(f'<x id="{marker_id}"/>')
        offset = match.end()

    parts.append(html.escape(text[offset:]))
    return "".join(parts), tokens


def restore_markup(text: str, tokens: list[str]) -> str:
    seen: list[int] = []

    def replace_marker(match: re.Match[str]) -> str:
        marker_id = int(match.group(1))
        if marker_id >= len(tokens):
            raise ValueError(f"DeepL returned an unknown markup marker: {marker_id}")
        seen.append(marker_id)
        return tokens[marker_id]

    restored = DEEPL_MARKER_RE.sub(replace_marker, text)
    if sorted(seen) != list(range(len(tokens))):
        raise ValueError(
            f"DeepL changed protected markup markers: expected {len(tokens)}, got {seen}"
        )
    return html.unescape(restored)


def join_dialog_source(texts: list[str], source_lang: str) -> str:
    """Undo the engine's visual wrapping before sending a dialog to DeepL."""
    separator = "" if source_lang.upper() in {"JA", "ZH"} else " "
    return separator.join(text.strip() for text in texts)


def collapse_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def choose_line_boundaries(
    pieces: list[str], source_lines: list[str], glue: int
) -> list[int]:
    """The indexes that cut `pieces` into one run per source line.

    `glue` is how many characters end up between two pieces once they are
    joined back, so that the target lengths are measured against the text the
    reader will see rather than against the pieces alone.
    """
    line_count = len(source_lines)
    source_lengths = [max(1, len(line.strip())) for line in source_lines]
    source_total = sum(source_lengths)
    rendered_total = sum(len(piece) for piece in pieces) + glue * max(
        0, len(pieces) - 1
    )
    prefix_lengths = [0]
    for index, piece in enumerate(pieces, start=1):
        prefix_lengths.append(
            prefix_lengths[-1] + len(piece) + (glue if index > 1 else 0)
        )

    boundaries = [0]
    cumulative_source = 0
    for line_index in range(line_count - 1):
        cumulative_source += source_lengths[line_index]
        target = rendered_total * cumulative_source / source_total
        minimum = boundaries[-1] + 1
        maximum = len(pieces) - (line_count - line_index - 1)
        if minimum > maximum:
            boundaries.append(min(len(pieces), boundaries[-1] + 1))
            continue
        boundary = min(
            range(minimum, maximum + 1),
            key=lambda candidate: abs(prefix_lengths[candidate] - target),
        )
        boundaries.append(boundary)
    boundaries.append(len(pieces))
    return boundaries


def split_dialog_translation(translated: str, source_lines: list[str]) -> list[str]:
    """Wrap a translation proportionally, without cutting through words.

    Spaces are where a translation is wrapped when it has enough of them. A
    window of moaning does not: DeepL answers a line like 「…ひ…ぎ…っ…」 with one
    unbroken run chained by ellipses, and a window of three lines then has a
    single word to spread over them. The tail lines used to come out empty, and
    an empty cell is dropped when the overlay is built - the engine draws its
    Japanese instead, so the window changed language halfway through. Falling
    back to punctuation gives such a run the boundaries it lacks. Windows that
    have a word per line are split exactly as before.
    """
    line_count = len(source_lines)
    if line_count <= 1:
        return [collapse_whitespace(translated)]

    token_pattern = re.compile(PROTECTED_TOKEN_RE.pattern + r"|\S+")
    words = token_pattern.findall(translated)
    if not words:
        return [""] * line_count

    if len(words) >= line_count:
        pieces, glue = words, 1
    else:
        # An atom keeps the punctuation it ends on, so the atoms of a text
        # concatenate back into it and nothing is glued between them.
        atoms = DIALOG_ATOM_RE.findall(translated)
        pieces, glue = (atoms, 0) if len(atoms) > len(words) else (words, 1)

    boundaries = choose_line_boundaries(pieces, source_lines, glue)
    separator = " " if glue else ""
    lines = [
        separator.join(pieces[boundaries[index] : boundaries[index + 1]])
        for index in range(line_count)
    ]
    return lines if glue else [collapse_whitespace(line) for line in lines]


def http_error_message(exc: urllib.error.HTTPError) -> str:
    try:
        body = exc.read().decode("utf-8", errors="replace").strip()
    except OSError:
        body = ""

    if body:
        try:
            parsed = json.loads(body)
            body = str(parsed.get("message", body))
        except json.JSONDecodeError:
            pass
    return body or str(exc.reason)


def request_deepl(
    api_key: str,
    api_url: str,
    body_fields: dict[str, Any],
    expected_count: int,
    retries: int,
    cancel_file: Path | None = None,
) -> list[str]:
    body = json.dumps(body_fields, ensure_ascii=False).encode("utf-8")

    for attempt in range(1, retries + 1):
        check_cancelled(cancel_file)
        request = urllib.request.Request(
            api_url,
            data=body,
            headers={
                "Authorization": f"DeepL-Auth-Key {api_key}",
                "Content-Type": "application/json",
                "User-Agent": "delta-rsan-overlay-translator/1.0",
            },
            method="POST",
        )

        try:
            with urllib.request.urlopen(request, timeout=120) as response:
                parsed = json.loads(response.read().decode("utf-8"))
            translations = parsed.get("translations", [])
            if len(translations) != expected_count:
                raise ValueError(
                    f"DeepL returned {len(translations)} translations for "
                    f"{expected_count} texts"
                )
            return [str(item["text"]) for item in translations]
        except urllib.error.HTTPError as exc:
            message = http_error_message(exc)
            if exc.code == 403:
                raise RuntimeError(
                    "DeepL rejected the API key (HTTP 403). Check DEEPL_API_KEY and "
                    "whether --api-url matches your Free or paid account. "
                    f"DeepL message: {message}"
                ) from exc
            if exc.code == 456:
                raise RuntimeError(
                    "DeepL monthly character quota is exhausted (HTTP 456). "
                    f"DeepL message: {message}"
                ) from exc

            retryable = exc.code == 429 or 500 <= exc.code < 600
            if not retryable or attempt == retries:
                raise RuntimeError(
                    f"DeepL API error HTTP {exc.code}: {message}"
                ) from exc

            retry_after = exc.headers.get("Retry-After")
            sleep_for = (
                int(retry_after)
                if retry_after and retry_after.isdigit()
                else min(60, 2**attempt)
            )
            print(
                f"DeepL API error HTTP {exc.code}: {message}. "
                f"Retrying in {sleep_for}s..."
            )
            sleep_with_cancellation(sleep_for, cancel_file)
        except (
            urllib.error.URLError,
            TimeoutError,
            ValueError,
            json.JSONDecodeError,
        ) as exc:
            if attempt == retries:
                raise RuntimeError(f"DeepL request failed: {exc}") from exc
            sleep_for = min(60, 2**attempt)
            print(f"DeepL request failed: {exc}. Retrying in {sleep_for}s...")
            sleep_with_cancellation(sleep_for, cancel_file)

    raise RuntimeError("unreachable")


def check_cancelled(cancel_file: Path | None) -> None:
    if cancel_file is not None and cancel_file.exists():
        raise KeyboardInterrupt


def sleep_with_cancellation(seconds: int, cancel_file: Path | None) -> None:
    deadline = time.monotonic() + seconds
    while True:
        check_cancelled(cancel_file)
        remaining = deadline - time.monotonic()
        if remaining <= 0:
            return
        time.sleep(min(0.2, remaining))


def call_deepl(
    api_key: str,
    api_url: str,
    source_lang: str,
    target_lang: str,
    texts: list[str],
    retries: int,
    cancel_file: Path | None = None,
) -> list[str]:
    protected = [protect_markup(text) for text in texts]
    context = html.escape(
        XML_INVALID_CHAR_RE.sub("\uFFFD", "\n".join(texts)), quote=False
    )
    for index, (fragment, _) in enumerate(protected):
        try:
            ET.fromstring(f"<root>{fragment}</root>")
        except ET.ParseError as exc:
            raise ValueError(
                f"Source text {index + 1} is not valid XML after escaping: {exc}"
            ) from exc

    translated = request_deepl(
        api_key,
        api_url,
        {
            "text": [item[0] for item in protected],
            "source_lang": source_lang.upper(),
            "target_lang": target_lang.upper(),
            "context": context,
            "preserve_formatting": True,
            "split_sentences": "nonewlines",
            "tag_handling": "xml",
            "tag_handling_version": "v2",
            "ignore_tags": ["x"],
        },
        len(texts),
        retries,
        cancel_file,
    )
    return [
        restore_markup(text, protected[index][1])
        for index, text in enumerate(translated)
    ]


def dialog_translation_loop_reason(
    translated: str, source_lines: list[str], source_lang: str
) -> str | None:
    """Names the machine-loop signature in an answer, or returns None.

    Repetition by itself is valid in this script: cries, stammers and SFX often
    repeat the same short word many times. A retry is therefore allowed only
    when the answer is both long and greatly inflated relative to its source,
    and contains either an implausible character run or a consecutive repeated
    word sequence. This deliberately prefers missing a borderline case over
    changing an intentional repetition.
    """
    text = collapse_whitespace(translated)
    source = join_dialog_source(source_lines, source_lang)
    if len(text) < 120 or len(text) < max(1, len(source)) * 2.5:
        return None

    if re.search(r"(.)\1{19,}", text, re.DOTALL):
        return "character-run"

    words = [match.group(0).casefold() for match in LOOP_WORD_RE.finditer(text)]
    for width in range(1, min(5, len(words) // 6) + 1):
        for start in range(len(words) - width * 6 + 1):
            unit = words[start : start + width]
            repeats = 1
            while words[
                start + repeats * width : start + (repeats + 1) * width
            ] == unit:
                repeats += 1
            repeated_letters = repeats * sum(len(word) for word in unit)
            if repeats >= 6 and repeated_letters >= 60:
                return "word-sequence"
    return None


def dialog_translation_looks_looped(
    translated: str, source_lines: list[str], source_lang: str
) -> bool:
    """Whether an answer has a conservative machine-loop signature."""
    return dialog_translation_loop_reason(translated, source_lines, source_lang) is not None


def unresolved_loop_records(
    jobs: list[DialogJob],
    translated_dialogs: list[list[str]],
    source_lang: str,
) -> list[dict[str, Any]]:
    """Detailed unresolved loops from one completed API batch."""
    records: list[dict[str, Any]] = []
    for job, translated_lines in zip(jobs, translated_dialogs):
        if job.key.startswith("C"):
            continue
        source_lines = [row.source for row in job.rows]
        reason = dialog_translation_loop_reason(
            " ".join(translated_lines), source_lines, source_lang
        )
        if reason is None:
            continue
        records.append(
            {
                "dialog": job.key,
                "reason": reason,
                "source_lines": source_lines,
                "translation_lines": translated_lines,
                "source_characters": len(join_dialog_source(source_lines, source_lang)),
                "translation_characters": len(
                    collapse_whitespace(" ".join(translated_lines))
                ),
            }
        )
    return records


def update_unresolved_loop_report(
    path: Path,
    target_lang: str,
    processed_dialogs: set[str],
    unresolved: list[dict[str, Any]],
    input_xlsx: Path,
    output_xlsx: Path,
) -> int:
    """Merges this run into this language's per-game loop report."""
    language = target_lang.upper()
    report: dict[str, Any] = {"version": 1, "language": language}
    if path.exists():
        loaded = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(loaded, dict):
            # Migrate the short-lived combined format if one is encountered.
            if isinstance(loaded.get("languages"), dict):
                report.update(loaded["languages"].get(language, {}))
            else:
                report.update(loaded)
    report["version"] = 1
    report["language"] = language
    previous = report
    by_dialog = {
        str(item["dialog"]): item
        for item in previous.get("dialogs", [])
        if isinstance(item, dict) and "dialog" in item
    }
    for dialog in processed_dialogs:
        by_dialog.pop(dialog, None)
    for item in unresolved:
        by_dialog[str(item["dialog"])] = item

    def dialog_sort_key(item: dict[str, Any]) -> tuple[str, int]:
        dialog = str(item["dialog"])
        number = dialog[1:]
        return dialog[:1], int(number) if number.isdigit() else 0

    dialogs = sorted(by_dialog.values(), key=dialog_sort_key)
    updated_at = datetime.now(timezone.utc).isoformat()
    report["version"] = 1
    report["language"] = language
    report["updated_at"] = updated_at
    report["input_workbook"] = str(input_xlsx.resolve())
    report["output_workbook"] = str(output_xlsx.resolve())
    report["unresolved_count"] = len(dialogs)
    report["dialogs"] = dialogs
    report.pop("languages", None)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    temporary.replace(path)
    return len(dialogs)


def request_dialog_texts(
    api_key: str,
    api_url: str,
    source_lang: str,
    target_lang: str,
    texts: list[str],
    retries: int,
    cancel_file: Path | None,
) -> list[str]:
    """Sends dialog texts with markup protected and restores the answers."""
    protected = [protect_markup(text) for text in texts]
    translated = request_deepl(
        api_key,
        api_url,
        {
            "text": [item[0] for item in protected],
            "source_lang": source_lang.upper(),
            "target_lang": target_lang.upper(),
            "preserve_formatting": True,
            "split_sentences": "nonewlines",
            "tag_handling": "xml",
            "tag_handling_version": "v2",
            "ignore_tags": ["x"],
        },
        len(texts),
        retries,
        cancel_file,
    )
    return [
        restore_markup(text, protected[index][1])
        for index, text in enumerate(translated)
    ]


def call_deepl_dialogs(
    api_key: str,
    api_url: str,
    source_lang: str,
    target_lang: str,
    dialogs: list[list[str]],
    retries: int,
    cancel_file: Path | None = None,
    per_line: list[bool] | None = None,
    loop_retries: int = DEFAULT_LOOP_RETRIES,
) -> list[list[str]]:
    """Translates each dialog and returns its lines, one per source line.

    `per_line[index]` decides how. False is a message window: its lines are one
    sentence the engine wrapped, so they go as a single text and the answer is
    wrapped back over the same number of lines. True is a menu: every line is an
    option in its own right and gets its own text, which is the only way each
    one comes back a complete phrase. With `loop_retries` above zero, a
    pathologically looped joined answer is retried as separate lines; this
    opt-in recovery never applies to menus.
    """
    if per_line is None:
        per_line = [False] * len(dialogs)

    texts: list[str] = []
    spans: list[tuple[int, int]] = []
    for lines, alone in zip(dialogs, per_line):
        start = len(texts)
        texts.extend(lines if alone else [join_dialog_source(lines, source_lang)])
        spans.append((start, len(texts)))

    restored = request_dialog_texts(
        api_key,
        api_url,
        source_lang,
        target_lang,
        texts,
        retries,
        cancel_file,
    )
    results: list[list[str]] = []
    for (start, end), lines, alone in zip(spans, dialogs, per_line):
        if alone:
            results.append(
                [collapse_whitespace(piece) for piece in restored[start:end]]
            )
            continue

        joined_answer = restored[start]
        if loop_retries > 0 and dialog_translation_looks_looped(
            joined_answer, lines, source_lang
        ):
            retried: list[str] = []
            for loop_attempt in range(1, loop_retries + 1):
                print(
                    "DeepL returned a looped dialog; retrying its "
                    f"{len(lines)} source line(s) individually "
                    f"({loop_attempt}/{loop_retries})."
                )
                retried = request_dialog_texts(
                    api_key,
                    api_url,
                    source_lang,
                    target_lang,
                    lines,
                    retries,
                    cancel_file,
                )
                if not dialog_translation_looks_looped(
                    " ".join(retried), lines, source_lang
                ):
                    break
            results.append([collapse_whitespace(piece) for piece in retried])
        else:
            results.append(split_dialog_translation(joined_answer, lines))
    return results


def parse_dialog_refs(value: str) -> list[tuple[str, int, int]]:
    return [
        (match.group(1), int(match.group(2)), int(match.group(3)))
        for match in DIALOG_REF_RE.finditer(value)
    ]


def owner_dialog_id(row: WorkbookRow) -> str:
    """The D/C window whose contextual translation owns this physical row."""
    refs = parse_dialog_refs(row.dialog_line)
    if not refs:
        raise ValueError(
            f"Workbook row {row.row_index} has no valid Dialog Line; run Extract again"
        )
    # A source shared by a choice and a message must remain a self-contained
    # option. This is the same ownership rule build_dialog_jobs has always used.
    return next((key for key, _, _ in refs if key.startswith("C")), refs[0][0])


def collect_rows(
    workbook_path: Path,
    source_col_name: str,
    target_col_name: str,
    overwrite: bool,
) -> tuple[Any, list[WorkbookRow], list[WorkbookRow]]:
    wb = openpyxl.load_workbook(workbook_path)
    if "Scenario" not in wb.sheetnames:
        raise ValueError("Workbook does not have the Scenario sheet; run Extract again")
    ws = wb["Scenario"]
    source_col = find_header_cell(ws, source_col_name)
    target_col = find_header_cell(ws, target_col_name)
    try:
        dialog_col = find_header_cell(ws, "Dialog Line")
    except ValueError as exc:
        raise ValueError(
            "Workbook has no Dialog Line column; run Extract again"
        ) from exc
    all_rows: list[WorkbookRow] = []
    pending: list[WorkbookRow] = []

    for row_idx in range(2, ws.max_row + 1):
        source = cell_text(ws.cell(row_idx, source_col).value)
        if not source.strip():
            continue
        row = WorkbookRow(
            source=source,
            worksheet=ws,
            row_index=row_idx,
            target_column=target_col,
            dialog_line=cell_text(ws.cell(row_idx, dialog_col).value).strip(),
        )
        all_rows.append(row)
        target = cell_text(ws.cell(row_idx, target_col).value).strip()
        if overwrite or not target:
            pending.append(row)

    return wb, all_rows, pending


def select_dialog_jobs(
    all_rows: list[WorkbookRow],
    pending: list[WorkbookRow],
    cache: dict[str, Any],
    max_dialogs: int,
) -> tuple[list[WorkbookRow], list[WorkbookRow], list[DialogJob]]:
    uncached = [row for row in pending if owner_dialog_id(row) not in cache]
    jobs = build_dialog_jobs(all_rows, uncached, cache)
    selected_jobs = jobs[:max_dialogs] if max_dialogs else jobs
    selected_sources = {
        source for job in selected_jobs for source in job.wanted_sources
    }
    selected_rows = [row for row in uncached if row.source in selected_sources]
    return uncached, selected_rows, selected_jobs


def collect_windows(
    all_rows: list[WorkbookRow], require_complete: bool = True
) -> dict[str, list[WorkbookRow]]:
    """Every window in the workbook, as its rows in reading order.

    A row can sit in several windows: the workbook holds one row per distinct
    source line, and the script reuses lines. That is also why a window is a
    list of rows rather than of texts.

    Joining a window needs every one of its lines, so a gap is an error by
    default. Deciding which lines stand on their own does not, and is asked of
    partial row sets, so that caller turns the check off.
    """
    positions: OrderedDict[str, dict[int, WorkbookRow]] = OrderedDict()
    totals: dict[str, int] = {}

    for row in all_rows:
        refs = parse_dialog_refs(row.dialog_line)
        if not refs:
            raise ValueError(
                f"Workbook row {row.row_index} has no valid Dialog Line; run Extract again"
            )
        for dialog_id, position, total in refs:
            if position < 1 or total < 1 or position > total:
                raise ValueError(
                    f"Invalid Dialog Line at workbook row {row.row_index}: {row.dialog_line}"
                )
            previous_total = totals.setdefault(dialog_id, total)
            if previous_total != total:
                raise ValueError(f"Dialog {dialog_id} has conflicting line totals")
            dialog_positions = positions.setdefault(dialog_id, {})
            previous_row = dialog_positions.get(position)
            if previous_row is not None and previous_row.source != row.source:
                raise ValueError(f"Dialog {dialog_id} has two different lines at {position}")
            dialog_positions[position] = row

    windows: dict[str, list[WorkbookRow]] = {}
    for dialog_id, dialog_positions in positions.items():
        total = totals[dialog_id]
        expected = set(range(1, total + 1))
        if require_complete and set(dialog_positions) != expected:
            missing = sorted(expected - set(dialog_positions))
            raise ValueError(f"Dialog {dialog_id} is missing line positions {missing}")
        windows[dialog_id] = [
            dialog_positions[index]
            for index in range(1, total + 1)
            if index in dialog_positions
        ]
    return windows


def standalone_windows(windows: dict[str, list[WorkbookRow]]) -> set[str]:
    """The windows whose lines are translated one at a time, never joined.

    Every menu is one: an option is a phrase in its own right. The rule then
    spreads, because the engine stores one string per line and the workbook one
    row per string - a line used as an option is the same string wherever else
    it appears. Joining it into a message would make the menu show a slice of
    somebody's sentence, so a window holding such a line is translated line by
    line too; that makes its own lines self-contained, which carries on to any
    window sharing one of them. Mixing the two inside one window is what is
    being avoided: a window has to read as one piece, and a self-contained
    sentence next to two slices of another one does not.

    Across the shipped script the spread stops immediately - the two shared
    lines are each the whole of the message window they appear in.
    """
    alone = {key for key in windows if key.startswith("C")}
    sources = {row.source for key in alone for row in windows[key]}
    while True:
        grown = {
            key
            for key, rows in windows.items()
            if key not in alone and any(row.source in sources for row in rows)
        }
        if not grown:
            return alone
        alone |= grown
        sources |= {row.source for key in grown for row in windows[key]}


def standalone_sources(windows: dict[str, list[WorkbookRow]]) -> set[str]:
    """The source lines that translate as texts of their own."""
    return {
        row.source for key in standalone_windows(windows) for row in windows[key]
    }


def build_dialog_jobs(
    all_rows: list[WorkbookRow],
    pending: list[WorkbookRow],
    cache: dict[str, Any],
) -> list[DialogJob]:
    windows = collect_windows(all_rows)
    alone = standalone_windows(windows)

    jobs: OrderedDict[str, DialogJob] = OrderedDict()
    for row in pending:
        refs = parse_dialog_refs(row.dialog_line)
        # A row belonging to both a menu and a message is translated with the
        # menu, which names the job; both windows are sent line by line anyway,
        # so the row gets the same text either way and the C label is the one
        # worth printing.
        dialog_id = owner_dialog_id(row)
        if dialog_id in cache:
            continue
        job = jobs.get(dialog_id)
        if job is None:
            job = DialogJob(
                dialog_id, windows[dialog_id], set(), dialog_id in alone
            )
            jobs[dialog_id] = job
        job.wanted_sources.add(row.source)
    return list(jobs.values())


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Translate a Delta RSAN.SD overlay workbook with the DeepL API."
    )
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    parser.add_argument("--source-col", default="Original")
    parser.add_argument("--target-col", default="TL")
    parser.add_argument("--source-lang", default="JA")
    parser.add_argument("--target-lang", default="RU")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--max-chars", type=int, default=25000)
    parser.add_argument("--max-items", type=int, default=DEEPL_MAX_TEXTS)
    parser.add_argument("--retries", type=int, default=5)
    parser.add_argument(
        "--loop-retries",
        type=int,
        default=DEFAULT_LOOP_RETRIES,
        help=(
            "Retry a detected DeepL content loop this many times, sending the "
            "window one source line per text; disabled by default"
        ),
    )
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument(
        "--max-dialogs",
        type=int,
        default=0,
        help=(
            "Translate at most this many uncached dialog windows; zero means all "
            "remaining windows"
        ),
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--api-key-stdin", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--cancel-file", type=Path, help=argparse.SUPPRESS)
    parser.add_argument(
        "--estimate",
        action="store_true",
        help=(
            "Report characters in complete uncached dialog windows without "
            "calling DeepL or writing output"
        ),
    )
    parser.add_argument(
        "--cache",
        type=Path,
        help="Translation cache; defaults to deepl_cache.jsonl beside the workbook",
    )
    args = parser.parse_args()
    if args.cache is None:
        args.cache = args.input_xlsx.resolve().parent / "deepl_cache.jsonl"

    if args.max_chars <= 0:
        parser.error("--max-chars must be greater than zero")
    if not 1 <= args.max_items <= DEEPL_MAX_TEXTS:
        parser.error(f"--max-items must be between 1 and {DEEPL_MAX_TEXTS}")
    if args.max_dialogs < 0:
        parser.error("--max-dialogs cannot be negative")
    if args.loop_retries < 0:
        parser.error("--loop-retries cannot be negative")

    wb, all_rows, pending = collect_rows(
        args.input_xlsx, args.source_col, args.target_col, args.overwrite
    )
    # Build ownership for the whole workbook, not only its empty rows. That lets
    # a normal no-op Translate run migrate a complete legacy source cache to the
    # contextual D/C format without spending API characters.
    all_jobs = build_dialog_jobs(all_rows, all_rows, {})
    if args.overwrite:
        cache: dict[str, tuple[str, ...]] = {}
        migrated: set[str] = set()
    else:
        cache, migrated = cache_for_dialog_jobs(
            load_dialog_cache(args.cache, args.target_lang),
            load_cache(args.cache, args.target_lang),
            all_jobs,
        )
    pending_jobs = build_dialog_jobs(all_rows, pending, {})
    uncached_rows, selected_rows, jobs = select_dialog_jobs(
        all_rows, pending, cache, args.max_dialogs
    )
    api_key = (
        sys.stdin.readline().rstrip("\r\n")
        if args.api_key_stdin
        else os.environ.get("DEEPL_API_KEY")
    )
    batches = iter_batches(jobs, args.max_chars, args.max_items)

    print(f"Workbook: {args.input_xlsx}")
    print(f"Uncached source texts: {len(uncached_rows)}")
    print(f"Selected source texts: {len(selected_rows)}")
    print(f"Dialog windows sent: {len(jobs)}")
    alone = [job for job in jobs if job.per_line]
    if alone:
        menus = sum(1 for job in alone if job.key.startswith("C"))
        print(
            f"  of which sent one line at a time: {len(alone)} "
            f"({menus} menus, {len(alone) - menus} windows sharing a line with one)"
        )
    print(f"Source characters: {sum(len(item.source) for item in selected_rows)}")
    print(f"Estimated API characters: {sum(job.character_count for job in jobs)}")
    if args.estimate:
        print("Estimate only: no API request sent")
        return
    print(f"Batches: {len(batches)}")
    if args.dry_run:
        return

    if migrated:
        jobs_by_key = {job.key: job for job in all_jobs}
        for key in sorted(migrated):
            append_dialog_cache(
                args.cache, jobs_by_key[key], cache[key], args.target_lang
            )
        print(f"Migrated legacy cache windows: {len(migrated)}")

    pending_by_source = {row.source: row for row in pending}
    done = 0
    processed_dialogs: set[str] = set()
    unresolved: list[dict[str, Any]] = []
    for job in pending_jobs:
        translated_lines = cache.get(job.key)
        if translated_lines is None:
            continue
        written: set[str] = set()
        for source_row, translated in zip(job.rows, translated_lines):
            if source_row.source not in job.wanted_sources:
                continue
            if source_row.source in written:
                continue
            written.add(source_row.source)
            pending_row = pending_by_source[source_row.source]
            pending_row.worksheet.cell(
                pending_row.row_index, pending_row.target_column
            ).value = translated
            done += 1

    try:
        if jobs and not api_key:
            raise SystemExit(
                "Set DEEPL_API_KEY before running this script "
                f"({len(selected_rows)} selected uncached source texts)."
            )
        for batch_idx, batch in enumerate(batches, start=1):
            check_cancelled(args.cancel_file)
            translated_dialogs = call_deepl_dialogs(
                api_key=api_key or "",
                api_url=args.api_url,
                source_lang=args.source_lang,
                target_lang=args.target_lang,
                dialogs=[[row.source for row in job.rows] for job in batch],
                retries=args.retries,
                cancel_file=args.cancel_file,
                per_line=[job.per_line for job in batch],
                loop_retries=args.loop_retries,
            )
            processed_dialogs.update(job.key for job in batch)
            unresolved.extend(
                unresolved_loop_records(batch, translated_dialogs, args.source_lang)
            )
            for job, translated_lines in zip(batch, translated_dialogs):
                # A window can repeat one source line at two positions, and the
                # workbook has a single row for it. Taking only the first of the
                # split pieces keeps one translation per row instead of writing
                # the cell twice and appending two conflicting cache lines.
                written: set[str] = set()
                for source_row, translated in zip(job.rows, translated_lines):
                    if source_row.source not in job.wanted_sources:
                        continue
                    if source_row.source in written:
                        continue
                    written.add(source_row.source)
                    pending_row = pending_by_source[source_row.source]
                    pending_row.worksheet.cell(
                        pending_row.row_index, pending_row.target_column
                    ).value = translated
                    done += 1

                cache[job.key] = tuple(translated_lines)
                append_dialog_cache(
                    args.cache, job, translated_lines, args.target_lang
                )

            print(f"Batch {batch_idx}/{len(batches)} done. Filled rows: {done}")
            check_cancelled(args.cancel_file)
    finally:
        args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb.save(args.output_xlsx)
        print(f"Workbook saved. Filled rows: {done}")
        report_path = args.input_xlsx.resolve().parent / unresolved_loop_report_name(
            args.target_lang
        )
        unresolved_count = update_unresolved_loop_report(
            report_path,
            args.target_lang,
            processed_dialogs,
            unresolved,
            args.input_xlsx,
            args.output_xlsx,
        )
        print(f"Unresolved loop report: {report_path}")
        print(
            f"Unresolved looped translations ({args.target_lang.upper()}): "
            f"{unresolved_count}"
        )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("Interrupted")
    except (OSError, ValueError, RuntimeError, KeyError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)
