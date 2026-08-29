"""Profile-driven workbook proofreader for Delta games.

The engine is deliberately VN-agnostic. A JSON rules file can provide simple
replacements, regex replacements, and row-specific fixes without changing this
script for every new game.

The rules file is per target language - `proofread_rules.ru.json` beside
`proofread_rules.en.json`. Almost nothing a proofreader does survives a change
of language: a replacement spells a name the way one language spells it, and a
`sources` entry is a whole translated line keyed by the Japanese original, which
is the same line in every workbook. Pointing the Russian rules at an English
workbook would overwrite English cells with Russian text.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install openpyxl"
    ) from exc

from delta_deepl import dialog_translation_looks_looped


# D is a message window, C a choice menu. A dialogs rule may name either, and a
# signature rule matches on the source lines, so relabelling a window leaves the
# rules that point at it working.
DIALOG_REF_RE = re.compile(r"\b([CD]\d+)\s+(\d+)/(\d+)\b")


def text(value: Any) -> str:
    return "" if value is None else str(value)


def headers(sheet: Any) -> dict[str, int]:
    return {
        text(sheet.cell(1, column).value).strip(): column
        for column in range(1, sheet.max_column + 1)
        if text(sheet.cell(1, column).value).strip()
    }


def load_rules(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    # utf-8-sig, not utf-8: the GUI opens this file in whatever the system has
    # registered for .json, and Notepad writes a BOM. json.load rejects one.
    with path.open("r", encoding="utf-8-sig") as handle:
        value = json.load(handle)
    if not isinstance(value, dict):
        raise ValueError("Proofread rules must be a JSON object")
    return value


def merge_rules(base: dict[str, Any], project: dict[str, Any]) -> dict[str, Any]:
    result = dict(base)
    for key, value in project.items():
        existing = result.get(key)
        if isinstance(existing, dict) and isinstance(value, dict):
            result[key] = merge_rules(existing, value)
        elif isinstance(existing, list) and isinstance(value, list):
            result[key] = existing + value
        else:
            result[key] = value
    return result


def apply_regex_rules(items: Any, text: str) -> str:
    result = text
    for item in items or []:
        if not isinstance(item, dict) or "pattern" not in item:
            raise ValueError("Each regex replacement needs pattern and replacement")
        result = re.sub(
            str(item["pattern"]),
            str(item.get("replacement", "")),
            result,
            flags=re.IGNORECASE if item.get("ignore_case") else 0,
        )
    return result


def apply_rule_block(source: str, current: str, rules: dict[str, Any]) -> str:
    result = current
    for old, new in dict(rules.get("replacements", {})).items():
        result = result.replace(str(old), str(new))

    result = apply_regex_rules(rules.get("regex_replacements"), result)

    for item in rules.get("source_rules", []):
        if not isinstance(item, dict):
            raise ValueError("Each source rule must be an object")
        source_pattern = item.get("source_pattern")
        if source_pattern is not None:
            try:
                matches_source = re.search(str(source_pattern), source, flags=re.DOTALL) is not None
            except re.error as exc:
                raise ValueError(f"Invalid source rule pattern: {source_pattern}") from exc
        else:
            matches_source = str(item.get("contains", "")) in source
        if not matches_source:
            continue
        for old, new in dict(item.get("replacements", {})).items():
            result = result.replace(str(old), str(new))
        result = apply_regex_rules(item.get("regex_replacements"), result)
    return result


def workbook_dialogs(
    sheet: Any, columns: dict[str, int]
) -> tuple[
    dict[str, tuple[tuple[int, str], ...]],
    dict[tuple[str, int], int],
]:
    """Returns complete windows as ordered ``(workbook row, source)`` pairs."""
    if "Dialog Line" not in columns:
        return {}, {}

    parts: dict[str, dict[int, tuple[int, int, str]]] = {}
    references: dict[tuple[str, int], int] = {}
    for row in range(2, sheet.max_row + 1):
        source = text(sheet.cell(row, columns["Original"]).value)
        dialog_cell = text(sheet.cell(row, columns["Dialog Line"]).value)
        for dialog_id, position_text, total_text in DIALOG_REF_RE.findall(dialog_cell):
            position = int(position_text)
            total = int(total_text)
            value = (total, row, source)
            previous = parts.setdefault(dialog_id, {}).get(position)
            if previous is not None and previous != value:
                raise ValueError(
                    f"Dialog {dialog_id} position {position} points to multiple rows"
                )
            parts[dialog_id][position] = value
            references[(dialog_id, position)] = row

    result: dict[str, tuple[tuple[int, str], ...]] = {}
    for dialog_id, positions in parts.items():
        totals = {value[0] for value in positions.values()}
        if len(totals) != 1:
            raise ValueError(f"Dialog {dialog_id} has inconsistent line totals")
        total = next(iter(totals))
        if set(positions) != set(range(1, total + 1)):
            continue
        result[dialog_id] = tuple(
            (positions[position][1], positions[position][2])
            for position in range(1, total + 1)
        )
    return result, references


def resolve_dialog_overrides(
    configured: Any,
    windows: dict[str, tuple[tuple[int, str], ...]],
    references: dict[tuple[str, int], int],
) -> tuple[dict[int, str], set[str], set[str]]:
    """Resolves legacy IDs and stable source-signature dialog rules to rows.

    A signature rule keeps its D-number only as a readable label. The ordered
    Japanese source lines are the identity, so inserting an earlier message
    cannot silently redirect a proofread fix to another scene.
    """
    if not isinstance(configured, dict):
        raise ValueError("dialogs must be an object keyed by Dialog Line ID")

    candidates: dict[int, list[tuple[str, str]]] = {}
    used: set[str] = set()
    missing: set[str] = set()
    for rule_id, value in configured.items():
        label = str(rule_id)
        if isinstance(value, list):
            translations = value
            found = False
            for position, translation in enumerate(translations, start=1):
                row = references.get((label, position))
                if row is None:
                    continue
                candidates.setdefault(row, []).append(
                    (f"{label} {position}", str(translation))
                )
                found = True
            if found:
                used.add(label)
            else:
                missing.add(label)
            continue
        elif isinstance(value, dict):
            originals = value.get("original")
            translations = value.get("translation")
            if not isinstance(originals, list) or not originals:
                raise ValueError(f"dialogs[{label!r}].original must be a non-empty array")
            if not isinstance(translations, list) or not translations:
                raise ValueError(
                    f"dialogs[{label!r}].translation must be a non-empty array"
                )
            signature = tuple(map(str, originals))
            matched_ids = [
                dialog_id
                for dialog_id, rows in windows.items()
                if tuple(source for _, source in rows) == signature
            ]
        else:
            raise ValueError(
                f"dialogs[{label!r}] must be an array or a signature object"
            )

        if not isinstance(translations, list) or not translations:
            raise ValueError(f"dialogs[{label!r}] must contain translations")
        if not matched_ids:
            if isinstance(value, dict):
                partial_matches = []
                for dialog_id, rows in windows.items():
                    sources = tuple(source for _, source in rows)
                    width = len(signature)
                    if any(
                        sources[index : index + width] == signature
                        for index in range(len(sources) - width + 1)
                    ):
                        partial_matches.append((dialog_id, sources))
                if partial_matches:
                    preview = "; ".join(
                        f"{dialog_id}: {list(sources)!r}"
                        for dialog_id, sources in partial_matches[:3]
                    )
                    remaining = len(partial_matches) - 3
                    if remaining > 0:
                        preview += f"; and {remaining} more"
                    raise ValueError(
                        f"Dialog signature {label} has no matching source window; "
                        f"it is now only part of expanded window(s) {preview}. "
                        "Update both original and translation arrays."
                    )
                raise ValueError(
                    f"Dialog signature {label} has no matching source window"
                )
            missing.add(label)
            continue
        for dialog_id in matched_ids:
            rows = windows[dialog_id]
            if len(translations) != len(rows):
                raise ValueError(
                    f"Dialog rule {label} has {len(translations)} translations, "
                    f"but matched window {dialog_id} has {len(rows)} lines"
                )
            for position, ((row, _), translation) in enumerate(
                zip(rows, translations), start=1
            ):
                candidates.setdefault(row, []).append(
                    (f"{label} {position}", str(translation))
                )
        used.add(label)

    overrides: dict[int, str] = {}
    for row, matches in candidates.items():
        values = {value for _, value in matches}
        if len(values) > 1:
            detail = "; ".join(f"{label} -> {value!r}" for label, value in matches)
            raise ValueError(
                f"Conflicting dialog rules at workbook row {row}: one shared row "
                f"cannot hold different text for {detail}"
            )
        overrides[row] = next(iter(values))
    return overrides, used, missing


def unresolved_looped_dialogs(
    sheet: Any,
    columns: dict[str, int],
    windows: dict[str, tuple[tuple[int, str], ...]],
) -> list[str]:
    """Finds looped translations that remain in the completed proofread book."""
    unresolved: list[str] = []
    for dialog_id, rows in windows.items():
        source_lines = [source for _, source in rows]
        translated_lines = [text(sheet.cell(row, columns["TL"]).value) for row, _ in rows]
        if not all(line.strip() for line in translated_lines):
            continue
        translated = " ".join(translated_lines)
        if dialog_translation_looks_looped(translated, source_lines, "JA"):
            unresolved.append(dialog_id)
    return unresolved




def proofread(
    input_path: Path,
    output_path: Path,
    rules_path: Path | None,
    backup_path: Path | None,
    profile_path: Path | None = None,
) -> None:
    rules = merge_rules(load_rules(profile_path), load_rules(rules_path))
    workbook = openpyxl.load_workbook(input_path)
    sheet = workbook["Scenario"] if "Scenario" in workbook.sheetnames else workbook.active
    columns = headers(sheet)
    if "Original" not in columns or "TL" not in columns:
        raise ValueError("Workbook must contain Original and TL columns")

    manual_rows = {
        str(key): str(value)
        for key, value in dict(rules.get("rows", {})).items()
    }
    manual_sources = {
        str(key): str(value)
        for key, value in dict(rules.get("sources", {})).items()
    }
    configured_dialogs = rules.get("dialogs", {})
    if configured_dialogs and "Dialog Line" not in columns:
        raise ValueError("Workbook must contain Dialog Line when dialogs rules are used")
    dialog_windows, dialog_references = workbook_dialogs(sheet, columns)
    dialog_overrides, _used_dialogs, missing_dialogs = resolve_dialog_overrides(
        configured_dialogs, dialog_windows, dialog_references
    )
    used_sources: set[str] = set()
    changed = 0
    for row in range(2, sheet.max_row + 1):
        source = text(sheet.cell(row, columns["Original"]).value)
        current = text(sheet.cell(row, columns["TL"]).value)
        updated = dialog_overrides.get(row)
        if updated is None:
            updated = manual_rows.get(str(row))
        if updated is None and source in manual_sources:
            updated = manual_sources[source]
            used_sources.add(source)
        if updated is None:
            updated = apply_rule_block(source, current, rules)
        if updated == current:
            continue
        sheet.cell(row, columns["TL"]).value = updated
        if "Notes" in columns:
            # The Notes column belongs to the reviewer; only fill it when empty.
            notes = sheet.cell(row, columns["Notes"])
            if not text(notes.value).strip():
                notes.value = "Profile proofread"
        changed += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if backup_path is not None:
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(input_path, backup_path)
    workbook.save(output_path)
    unresolved_loops = unresolved_looped_dialogs(sheet, columns, dialog_windows)
    missing = sorted(set(manual_sources) - used_sources)
    if missing:
        print(f"Source rules with no matching line: {len(missing)}")
        for item in missing[:5]:
            print(f"  {item}")
    if missing_dialogs:
        print(f"Dialog rules with no matching window: {len(missing_dialogs)}")
        for item in sorted(missing_dialogs)[:5]:
            print(f"  {item}")
    print(f"Changed rows: {changed}")
    print(f"Proofread unresolved looped translations: {len(unresolved_loops)}")
    if unresolved_loops:
        preview = ", ".join(unresolved_loops[:12])
        remaining = len(unresolved_loops) - 12
        if remaining > 0:
            preview += f", and {remaining} more"
        print(
            "WARNING: unresolved looped translations remain after proofread: "
            f"{len(unresolved_loops)} ({preview})",
            file=sys.stderr,
        )
    print(f"Reviewed workbook: {output_path.resolve()}")
    if backup_path is not None:
        print(f"Backup: {backup_path.resolve()}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Proofread a Delta translation workbook")
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--profile", type=Path)
    parser.add_argument("--rules", type=Path)
    parser.add_argument("--backup", type=Path)
    args = parser.parse_args()
    try:
        proofread(args.input, args.output, args.rules, args.backup, args.profile)
    except (OSError, ValueError, KeyError, json.JSONDecodeError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
