"""Delta/Route2 scenario text: extract RSAN.SD into a workbook, and compile a
workbook back into the runtime overlay winmm.dll reads.

The overlay carries its own byte-to-character table, because translated lines
are encoded in a private half-width codepage rather than the SJIS tunnel - see
the format notes in README.md.
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import re
import struct
import sys
import unicodedata
from collections import Counter, OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import delta_fit

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: "
        "python -m pip install -r requirements.txt"
    ) from exc


SUPPORTED_SD_SHA256 = "d3c8cc0fa52e2a0b92b7452cb78b4ec22fa2794501c720e9021a82a8a41de2f0"
OVERLAY_MAGIC = b"RKT3"
HEADERS = [
    "ID",
    "Offset",
    "Speaker",
    "Dialog Line",
    "Original",
    "TL",
    "Occurrences",
    "Notes",
    "SourceBytesBase64",
]
NAME_TRANSLATIONS = {
    "RU": {
        "羽山　卓": "Таку Хаяма",
        "犬": "Пёс",
        "犬Ａ": "Пёс А",
        "犬Ｂ": "Пёс Б",
        "犬Ｃ": "Пёс В",
        "犬Ａと犬Ｂ": "Псы А и Б",
        "犬たち": "Псы",
        "豚": "Свинья",
        "青年Ａ": "Юноша А",
        "青年Ｂ": "Юноша Б",
        "青年Ｃ": "Юноша В",
        "ホームレスＡ": "Бродяга А",
        "ホームレスＢ": "Бродяга Б",
        "ホームレスＣ": "Бродяга В",
        "ホームレスＤ": "Бродяга Г",
        "ホームレスＥ": "Бродяга Д",
        "ホームレスＦ": "Бродяга Е",
        "運転手": "Водитель",
        "麗佳": "Рейка",
        "二階堂麗佳": "Рейка Никайдо",
        "玲": "Рей",
        "宮内玲": "Рей Мияути",
        "羽山": "Хаяма",
        "山田": "Ямада",
        "朱神公一": "Коити Акагами",
        "公一": "Коити",
        "ジョン": "Джон",
        "メイド": "Горничная",
        "ガードマン": "Охранник",
        "男": "Мужчина",
        "女": "Женщина",
        "青年": "Молодой человек",
        "お嬢様": "Госпожа",
    },
    "EN": {
        "羽山　卓": "Taku Hayama",
        "犬": "Dog",
        "犬Ａ": "Dog A",
        "犬Ｂ": "Dog B",
        "犬Ｃ": "Dog C",
        "犬Ａと犬Ｂ": "Dogs A and B",
        "犬たち": "The dogs",
        "豚": "Pig",
        "青年Ａ": "Young man A",
        "青年Ｂ": "Young man B",
        "青年Ｃ": "Young man C",
        "ホームレスＡ": "Homeless A",
        "ホームレスＢ": "Homeless B",
        "ホームレスＣ": "Homeless C",
        "ホームレスＤ": "Homeless D",
        "ホームレスＥ": "Homeless E",
        "ホームレスＦ": "Homeless F",
        "運転手": "Driver",
        "麗佳": "Reika",
        "二階堂麗佳": "Reika Nikaido",
        "玲": "Rei",
        "宮内玲": "Rei Miyauchi",
        "羽山": "Hayama",
        "山田": "Yamada",
        "朱神公一": "Koichi Akagami",
        "公一": "Koichi",
        "ジョン": "John",
        "メイド": "Maid",
        "ガードマン": "Guard",
        "男": "Man",
        "女": "Woman",
        "青年": "Young man",
        "お嬢様": "Milady",
    },
}

# RSA.EXE draws the context menu directly with its text renderer. These
# captions are not part of RSAN.SD and therefore do not appear in the
# translation workbook. Keep them in the runtime overlay so the same hook
# localizes the right-click menu as well as scenario text.
RUNTIME_MENU_TRANSLATIONS = {
    "RU": {
        "自動送り": "Автопрокрутка",
        "枠消去": "Скрыть рамку",
        "スキップ": "Пропустить",
        "最初に戻る": "С начала",
        "メニューに戻る": "В главное меню",
        "クイックセーブ": "Быстрое сохранение",
        "クイックロード": "Быстрая загрузка",
        "セーブ": "Сохранить",
        "ロード": "Загрузить",
        "終了": "Выйти",
        "画面効果": "Эффекты экрана",
        "操作": "Управление",
        "音響効果": "Звуковые эффекты",
        "ヘルプ": "Справка",
        "Delta web": "Сайт Delta",
    },
    "EN": {
        "自動送り": "Auto-Advance",
        "枠消去": "Hide Frame",
        "スキップ": "Skip",
        "最初に戻る": "Restart",
        "メニューに戻る": "Return to Main Menu",
        "クイックセーブ": "Quick Save",
        "クイックロード": "Quick Load",
        "セーブ": "Save",
        "ロード": "Load",
        "終了": "Exit",
        "画面効果": "Visual Effects",
        "操作": "Controls",
        "音響効果": "Sound Effects",
        "ヘルプ": "Help",
        "Delta web": "Delta Website",
    },
}


# The name plate is drawn from the //NameStr table in MS.MHU, not from the
# scenario, and every entry is padded with ideographic spaces. Those padded
# strings need overlay keys of their own or the plate stays Japanese.
NAME_PLATES = [
    "　羽山　卓　",
    "　　麗佳　　",
    "　　玲　　　",
    "　　山田　　",
    "　ジョン　　",
    "　　犬　　　",
    "　　犬Ａ　　",
    "　　犬Ｂ　　",
    "　　犬Ｃ　　",
    "犬Ａと犬Ｂ　",
    " 　犬たち　 ",
    "　　豚　　　",
    "　　男　　　",
    "　　女　　　",
    "　青年Ａ　　",
    "　青年Ｂ　　",
    "　青年Ｃ　　",
    "　運転手　　",
    "ホームレスＡ",
    "ホームレスＢ",
    "ホームレスＣ",
    "ホームレスＤ",
    "ホームレスＥ",
    "ホームレスＦ",
    "ガードマン　",
]


# RSA.EXE advances the pen by fontHeight + 2 for a CP932 lead byte pair and by
# (fontHeight + 2) / 2 for every other byte, so encoding Cyrillic as single
# bytes halves the width of a translated line. NUL, tab and '@' are read by the
# renderer itself and must never appear in encoded text.
RESERVED_BYTES = frozenset({0x00, 0x09, 0x40})

RUSSIAN_ALPHABET = (
    "АБВГДЕЁЖЗИЙК"
    "ЛМНОПРСТУФХЦ"
    "ЧШЩЪЫЬЭЮЯ"
    "абвгдеёжзийк"
    "лмнопрстуфхц"
    "чшщъыьэюя"
)
EXTRA_GLYPHS = "«»…—"
MAPPED_CHARACTERS = RUSSIAN_ALPHABET + EXTRA_GLYPHS

# Typography DeepL emits that has no code of its own, folded onto something the
# encoder can represent.
NORMALIZATION = {
    "　": "  ",
    "\t": " ",
    " ": " ",
    "‒": "—",
    "–": "—",
    "―": "—",
    "─": "—",
    "“": "«",
    "”": "»",
    "„": "«",
    "‟": "«",
    "「": "«",
    "」": "»",
    "『": "«",
    "』": "»",
    "‘": "'",
    "’": "'",
    "（": "(",
    "）": ")",
    "！": "!",
    "，": ",",
    "．": ".",
    "：": ":",
    "；": ";",
    "？": "?",
}


def code_pool() -> list[int]:
    """Single-byte codes CP932 never treats as a lead byte."""
    pool = list(range(0xA0, 0xE0))
    pool += [0x80, 0xFD, 0xFE, 0xFF]
    pool += [ord(character) for character in "{}|^$<>#%&*+=~`"]
    return [code for code in pool if code not in RESERVED_BYTES]


def build_encoding() -> tuple[dict[str, int], dict[int, str]]:
    pool = code_pool()
    if len(MAPPED_CHARACTERS) > len(pool):
        raise ValueError(
            f"Overlay codepage needs {len(MAPPED_CHARACTERS)} codes, {len(pool)} available"
        )

    char_to_byte = {
        character: code for character, code in zip(MAPPED_CHARACTERS, pool)
    }
    custom_codes = set(char_to_byte.values())
    for code in range(0x20, 0x7F):
        if code in RESERVED_BYTES or code in custom_codes:
            continue
        char_to_byte.setdefault(chr(code), code)

    byte_to_char = {code: character for character, code in char_to_byte.items()}
    return char_to_byte, byte_to_char


def encode_translation(
    text: str, char_to_byte: dict[str, int], unmapped: Counter
) -> bytes:
    encoded = bytearray()
    for character in text:
        for piece in NORMALIZATION.get(character, character):
            code = char_to_byte.get(piece)
            if code is None:
                folded = unicodedata.normalize("NFKC", piece)
                if folded and all(part in char_to_byte for part in folded):
                    encoded.extend(char_to_byte[part] for part in folded)
                    continue
                unmapped[piece] += 1
                code = ord("?")
            encoded.append(code)
    return bytes(encoded)



if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")


@dataclass(frozen=True)
class TextEntry:
    source: bytes
    text: str
    offsets: tuple[int, ...]
    speakers: tuple[int, ...] = ()
    dialog_lines: tuple[str, ...] = ()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def is_japanese(text: str) -> bool:
    return any(
        "\u3040" <= char <= "\u30ff"
        or "\u3400" <= char <= "\u4dbf"
        or "\u4e00" <= char <= "\u9fff"
        or "\uf900" <= char <= "\ufaff"
        for char in text
    )


def has_kana(text: str) -> bool:
    """Full-width kana only.

    Japanese prose cannot be written without kana - particles, inflections and
    okurigana are all kana. Half-width katakana never appears in this script's
    text but turns up constantly as a stray byte inside binary records, so it
    is deliberately not counted as evidence of prose.
    """
    return any("\u3040" <= char <= "\u30ff" for char in text)


def is_full_width(char: str) -> bool:
    return (
        char == "\u3000"
        or "\u3001" <= char <= "\u303f"
        or "\u3040" <= char <= "\u30ff"
        or "\u3400" <= char <= "\u4dbf"
        or "\u4e00" <= char <= "\u9fff"
        or "\uf900" <= char <= "\ufaff"
        or "\uff01" <= char <= "\uff60"
        or char in "\u2010\u2014\u2015\u2018\u2019\u201c\u201d\u2026\u2032\u2033"
    )


def looks_like_scenario_text(text: str) -> bool:
    """Reject binary records that happen to decode as CP932.

    RSAN.SD interleaves text with byte structures. A three-byte record whose
    first byte lands in the CP932 lead range decodes to a plausible kanji - the
    extractor used to accept 121 of them, which cost DeepL characters and, worse,
    put junk lines inside real message windows: the dialog was then sent to the
    API with the junk prepended and the translation split across it, truncating
    the real line.

    Two ways to qualify. Text written entirely in full-width forms is scenario
    text whatever its length, which keeps kanji-only lines such as the short
    name calls. Text mixing in half-width or ASCII characters has to be long
    enough that those are punctuation rather than a stray byte: the shortest
    real line of that kind is seven characters, the longest junk record three.
    """
    if not is_japanese(text):
        return False
    if all(is_full_width(char) for char in text):
        return True
    return has_kana(text) and len(text) > 4


JAPANESE_QUOTE_PAIRS = {"「": "」", "『": "』"}


def has_complete_japanese_quote_frame(text: str) -> bool:
    """Whether text is enclosed by one matching pair of Japanese quotes."""
    return len(text) >= 2 and JAPANESE_QUOTE_PAIRS.get(text[0]) == text[-1]


def iter_null_chunks(data: bytes) -> Iterable[tuple[int, bytes]]:
    start = 0
    while start < len(data):
        end = data.find(b"\0", start)
        if end < 0:
            end = len(data)
        if end > start:
            yield start, data[start:end]
        start = end + 1


SPEAKER_COMMAND = re.compile(r"N(\d{1,2})\Z")
DIALOG_BOUNDARY_COMMANDS = {"M", "F"}

# A voiced line carries its clip in front of the text: the "vA######" sample
# name, then six bytes locating it. Nothing separates those six from the first
# character of the line, so whichever of them holds no zero stays in the same
# NUL-delimited chunk as the text - and the control-byte rule in extract_entries
# then discarded the whole line as a binary record. Reika is the only voiced
# speaker, so what was lost was her dialogue, roughly one line in five.
#
# The cut is by position rather than by looking for where prose starts: across
# the supported RSAN.SD every one of the 4089 voiced records puts its text
# exactly six bytes past the name terminator. Bytes before that are the locator
# whatever they decode to, which also keeps the ones that read as a lone kanji
# out of the workbook.
VOICE_COMMAND = re.compile(r"v[A-Za-z]\d{6}\Z")
VOICE_HEADER_SIZE = 6

# RSA.EXE draws a selectable list with opcode 0x49. The frame around it is
# closed on both sides, and that is what makes it safe to find by pattern
# instead of by interpreting the bytecode: the option count in the header has to
# be matched by exactly that many strings, and the run of strings has to land on
# the terminator. A binary record that happens to decode as CP932 can fake one
# end of that; it cannot fake both ends and the count between them.
#
# Whether the scenario acts on the answer is a second, independent test. A real
# player choice is followed by a jump table whose target count agrees with the
# header, so the number of options is stated twice, by two structures that have
# to say the same thing. The frames in the trailing code block of the supported
# script state it once: they carry no table and write story flags directly,
# which is a developer's state picker rather than a choice the reader is given.
CHOICE_OPEN = b"\x35\x00\x01\x00\x49\x00"
CHOICE_CLOSE = b"\x46\x00\xb2\x00\x04\x00\x35\x00\x00\x00"
CHOICE_BRANCH = b"\x92\x00\x04\x00"
CHOICE_MAX_OPTIONS = 15
CHOICE_MAX_OPTION_BYTES = 255


@dataclass(frozen=True)
class ChoiceBlock:
    """One 0x49 menu: where it starts, its options, and where it branches."""

    offset: int
    options: tuple[tuple[int, bytes], ...]
    targets: tuple[int, ...] = ()

    @property
    def branches(self) -> bool:
        """Whether a jump table consumes the answer, i.e. the reader picks it."""
        return bool(self.targets)


def read_choice_frame(data: bytes, offset: int) -> ChoiceBlock | None:
    """Reads the menu frame opening at `offset`, or None when it does not hold."""
    if data[offset:offset + len(CHOICE_OPEN)] != CHOICE_OPEN:
        return None
    header = offset + len(CHOICE_OPEN)
    if header + 2 > len(data):
        return None
    count = data[header]
    if not 1 <= count <= CHOICE_MAX_OPTIONS or data[header + 1] != 0:
        return None

    options: list[tuple[int, bytes]] = []
    cursor = header + 2
    for _ in range(count):
        end = data.find(b"\0", cursor)
        if end < 0 or end == cursor or end - cursor > CHOICE_MAX_OPTION_BYTES:
            return None
        option = data[cursor:end]
        # The same two rules the extractor applies to message text: a control
        # byte means the run has wandered out of the string table, and bytes the
        # engine could not render are not an option it could draw.
        if any(byte < 0x20 for byte in option):
            return None
        try:
            option.decode("cp932", errors="strict")
        except UnicodeDecodeError:
            return None
        options.append((cursor, option))
        cursor = end + 1

    if data[cursor:cursor + len(CHOICE_CLOSE)] != CHOICE_CLOSE:
        return None

    targets: tuple[int, ...] = ()
    table = cursor + len(CHOICE_CLOSE)
    if data[table:table + len(CHOICE_BRANCH)] == CHOICE_BRANCH:
        head = table + len(CHOICE_BRANCH)
        if head + 2 + 4 * count <= len(data) and data[head:head + 2] == bytes((count, 0)):
            addresses = tuple(
                struct.unpack_from("<I", data, head + 2 + 4 * index)[0]
                for index in range(count)
            )
            if all(0 < address < len(data) for address in addresses):
                targets = addresses

    return ChoiceBlock(offset, tuple(options), targets)


def find_choice_blocks(data: bytes) -> list[ChoiceBlock]:
    """Every 0x49 menu in the script, in file order."""
    blocks: list[ChoiceBlock] = []
    at = data.find(CHOICE_OPEN)
    while at >= 0:
        block = read_choice_frame(data, at)
        if block is not None:
            blocks.append(block)
        at = data.find(CHOICE_OPEN, at + 1)
    return blocks


def speaker_name(index: int) -> str:
    """N0 is narration; N1 and up select a plate from the //NameStr table, which
    is the same order NAME_PLATES was taken from. The plate is padded with
    ideographic spaces for the fixed-width box the engine draws it in, so it is
    trimmed here."""
    if index <= 0 or index > len(NAME_PLATES):
        return ""
    return NAME_PLATES[index - 1].replace("\u3000", "").strip()


def speaker_label(indices: tuple[int, ...]) -> str:
    names = [speaker_name(index) for index in indices]
    kept = list(dict.fromkeys(name for name in names if name))
    # A handful of short interjections are reused by two characters. Listing
    # both is honest; picking one would quietly mislabel the line.
    return " / ".join(kept)


def option_offsets(data: bytes) -> set[int]:
    """Where every option of every real choice sits in the script."""
    return {
        offset
        for block in find_choice_blocks(data)
        if block.branches
        for offset, _ in block.options
    }


def choice_windows(
    windows: list[list[tuple[int, bytes, str, int]]], options: set[int]
) -> set[int]:
    """Which windows hold choice text, by index into `windows`.

    A window holding an option holds choice text. So does a window that repeats
    one of those lines elsewhere in the script - the engine stores one string
    per line and the overlay replaces it by its bytes, so both places show the
    same words, and words that have to work as an option are written as a phrase
    that stands on its own. That property is what the C series records, and it
    has to spread until it stops: a window pulled in this way makes its own
    lines standalone in turn, which reaches any window sharing one of them.

    Without the spread a line could carry both a C and a D label, and the two
    disagree about how it is translated - an option is a text of its own, a
    message line is one slice of a sentence translated whole. Across the shipped
    script the spread adds two windows and stops there.
    """
    chosen = {
        index
        for index, lines in enumerate(windows)
        if any(offset in options for offset, _, _, _ in lines)
    }
    sources = {line[1] for index in chosen for line in windows[index]}
    while True:
        grown = {
            index
            for index, lines in enumerate(windows)
            if index not in chosen and any(line[1] in sources for line in lines)
        }
        if not grown:
            return chosen
        chosen |= grown
        sources |= {line[1] for index in grown for line in windows[index]}


def label_windows(
    windows: list[list[tuple[int, bytes, str, int]]], options: set[int]
) -> list[str]:
    """Names every window: C for choice text, D for a message.

    Both series are numbered in script order, and a choice window still consumes
    a message number. Numbering the C windows separately without doing that
    would renumber every message after the first menu, and proofread rules name
    windows.
    """
    chosen = choice_windows(windows, options)
    names: list[str] = []
    dialog_index = 0
    choice_index = 0
    for index in range(len(windows)):
        dialog_index += 1
        if index in chosen:
            choice_index += 1
            names.append(f"C{choice_index:05d}")
        else:
            names.append(f"D{dialog_index:05d}")
    return names


def extract_entries(path: Path) -> list[TextEntry]:
    data = path.read_bytes()
    grouped: OrderedDict[
        bytes, tuple[str, list[int], list[int], list[str]]
    ] = OrderedDict()

    # The script sets the name plate with an ASCII "N<index>" command and leaves
    # it in effect until the next one, so it has to be carried between strings.
    # M closes an ordinary message; F closes one before an effect or scene
    # transition. An N command after text also starts a new message. Across the
    # supported RSAN.SD this produces speaker-homogeneous windows of at most
    # three physical source lines.
    current = 0
    pending: list[tuple[int, bytes, str, int]] = []

    # Where the text of the voiced line being read begins; -1 between records.
    voiced_text_at = -1

    # Choice text gets its own C series so that a reviewer can tell a line the
    # reader picks from a line the reader is told, which is the difference
    # between writing an action and writing a fact. Numbering happens after the
    # whole script is read, because whether a window holds choice text depends
    # on windows further along - see label_windows.
    windows: list[list[tuple[int, bytes, str, int]]] = []

    def flush_dialog() -> None:
        if pending:
            windows.append(list(pending))
            pending.clear()

    for offset, chunk in iter_null_chunks(data):
        is_voiced_text = False
        if voiced_text_at >= 0:
            if offset + len(chunk) <= voiced_text_at:
                continue
            if offset < voiced_text_at:
                chunk = chunk[voiced_text_at - offset :]
                offset = voiced_text_at
            voiced_text_at = -1
            is_voiced_text = True

        try:
            command_text = chunk.decode("ascii", errors="strict")
        except UnicodeDecodeError:
            command_text = ""
        command = SPEAKER_COMMAND.fullmatch(command_text)
        if command is not None:
            flush_dialog()
            current = int(command.group(1))
            continue

        if command_text in DIALOG_BOUNDARY_COMMANDS:
            flush_dialog()
            continue

        if VOICE_COMMAND.fullmatch(command_text):
            voiced_text_at = offset + len(chunk) + 1 + VOICE_HEADER_SIZE
            continue

        # Tab used to be exempted here. Nothing in the scenario contains one;
        # every chunk that did was a binary record.
        if len(chunk) > 4096 or any(byte < 0x20 for byte in chunk):
            continue
        try:
            text = chunk.decode("cp932", errors="strict")
        except UnicodeDecodeError:
            continue
        if not looks_like_scenario_text(text) and not (
            is_voiced_text and has_complete_japanese_quote_frame(text)
        ):
            continue
        pending.append((offset, chunk, text, current))

    flush_dialog()

    for name, lines in zip(label_windows(windows, option_offsets(data)), windows):
        total = len(lines)
        for line_index, (offset, source, text, speaker) in enumerate(lines, start=1):
            location = f"{name} {line_index}/{total}"
            existing = grouped.get(source)
            if existing is None:
                grouped[source] = (text, [offset], [speaker], [location])
            else:
                existing[1].append(offset)
                if speaker not in existing[2]:
                    existing[2].append(speaker)
                existing[3].append(location)

    return [
        TextEntry(
            source=source,
            text=text,
            offsets=tuple(offsets),
            speakers=tuple(speakers),
            dialog_lines=tuple(dialog_lines),
        )
        for source, (text, offsets, speakers, dialog_lines) in grouped.items()
    ]


def verify_source(path: Path, allow_unknown: bool) -> str:
    digest = sha256(path)
    if digest != SUPPORTED_SD_SHA256 and not allow_unknown:
        raise ValueError(
            f"Unsupported RSAN.SD SHA-256: {digest}. Expected {SUPPORTED_SD_SHA256}. "
            "Use --allow-unknown only after checking the format manually."
        )
    return digest


def write_workbook(entries: list[TextEntry], output: Path, source_hash: str) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Scenario"
    sheet.append(HEADERS)

    for index, entry in enumerate(entries, start=1):
        sheet.append(
            [
                index,
                f"0x{entry.offsets[0]:06X}",
                speaker_label(entry.speakers),
                "; ".join(entry.dialog_lines),
                entry.text,
                "",
                len(entry.offsets),
                "",
                base64.b64encode(entry.source).decode("ascii"),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 8
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 26
    sheet.column_dimensions["E"].width = 72
    sheet.column_dimensions["F"].width = 72
    sheet.column_dimensions["G"].width = 13
    sheet.column_dimensions["H"].width = 28
    sheet.column_dimensions["I"].hidden = True
    for row in sheet.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[5].alignment = Alignment(wrap_text=True, vertical="top")

    metadata = workbook.create_sheet("Metadata")
    metadata.append(["Format", "Delta RSAN.SD runtime overlay"])
    metadata.append(["SourceSHA256", source_hash])
    metadata.append(["Encoding", "CP932 -> UTF-16LE/SJIS tunnel"])
    metadata.append(["DialogBoundaries", "M; F; N after text"])
    metadata.append(
        ["DialogLine", "D#### message window; C#### choice text, never mixed in a cell"]
    )
    metadata.sheet_state = "hidden"

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def find_column(sheet: object, name: str) -> int:
    max_column = sheet.max_column
    if max_column is None:
        header = next(
            sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
        )
        values = enumerate(header, start=1)
    else:
        values = (
            (column, sheet.cell(1, column).value)
            for column in range(1, max_column + 1)
        )
    for column, value in values:
        if str(value or "").strip().casefold() == name.casefold():
            return column
    raise ValueError(f"Missing XLSX column: {name}")


def workbook_translations(path: Path) -> list[tuple[bytes, str]]:
    workbook = openpyxl.load_workbook(path, data_only=False)
    sheet = workbook["Scenario"] if "Scenario" in workbook.sheetnames else workbook.active
    source_column = find_column(sheet, "SourceBytesBase64")
    tl_column = find_column(sheet, "TL")
    result: list[tuple[bytes, str]] = []

    for row in range(2, sheet.max_row + 1):
        translation = str(sheet.cell(row, tl_column).value or "").strip()
        encoded_source = str(sheet.cell(row, source_column).value or "").strip()
        if not translation or not encoded_source:
            continue
        source = base64.b64decode(encoded_source, validate=True)
        result.append((source, translation))
    return result


@dataclass(frozen=True)
class Coverage:
    """What the workbook holds, before anything is built from it.

    verify_source guards the input by hashing RSAN.SD. Nothing guarded the
    output: a workbook half-filled by an interrupted translation run built an
    overlay just as happily as a finished one, and the gap only showed up as
    Japanese still on screen.
    """

    rows: int
    translated: int
    windows: int
    complete_windows: int
    partial_windows: int
    empty_windows: int
    untranslated_samples: tuple[str, ...]

    @property
    def missing(self) -> int:
        return self.rows - self.translated

    @property
    def share(self) -> float:
        return self.translated / self.rows if self.rows else 0.0


def workbook_coverage(path: Path) -> Coverage:
    """Counts filled and empty translations, by row and by message window.

    A partially translated window is worse than an untranslated one: the reader
    gets a sentence that changes language halfway through. Counting windows as
    well as rows is what makes that visible.
    """
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        sheet = workbook["Scenario"] if "Scenario" in workbook.sheetnames else workbook.active
        tl_column = find_column(sheet, "TL")
        original_column = find_column(sheet, "Original")
        # Older workbooks predate this column, and proofread.py already treats
        # it as optional. Missing it costs the window figures, not the report.
        try:
            window_column = find_column(sheet, "Dialog Line")
        except ValueError:
            window_column = 0

        rows = 0
        translated = 0
        samples: list[str] = []
        # A row can sit in several windows, because rows are deduplicated by
        # source bytes and the same line recurs across the script.
        window_state: dict[str, list[int]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            original = str(row[original_column - 1] or "").strip()
            if not original:
                continue
            rows += 1
            filled = bool(str(row[tl_column - 1] or "").strip())
            translated += filled
            if not filled and len(samples) < 5:
                samples.append(original)
            if window_column:
                for location in str(row[window_column - 1] or "").split(";"):
                    identifier = location.split()
                    if identifier:
                        counts = window_state.setdefault(identifier[0], [0, 0])
                        counts[0] += 1
                        counts[1] += filled
    finally:
        workbook.close()

    complete = sum(1 for total, filled in window_state.values() if filled == total)
    empty = sum(1 for _, filled in window_state.values() if filled == 0)
    return Coverage(
        rows=rows,
        translated=translated,
        windows=len(window_state),
        complete_windows=complete,
        partial_windows=len(window_state) - complete - empty,
        empty_windows=empty,
        untranslated_samples=tuple(samples),
    )


def print_coverage(coverage: Coverage) -> None:
    print(
        f"Rows translated: {coverage.translated} of {coverage.rows} "
        f"({coverage.share:.1%})"
    )
    if coverage.windows:
        print(
            f"Windows: {coverage.complete_windows} complete, "
            f"{coverage.partial_windows} partial, {coverage.empty_windows} untranslated"
        )
    else:
        print("Windows: not counted, this workbook has no Dialog Line column")
    if coverage.partial_windows:
        print(
            "  warning: a partial window changes language mid-sentence on screen."
        )
    for sample in coverage.untranslated_samples:
        print(f"  untranslated: {sample[:60]}")


STANDALONE_FLAG = 1


def add_name_translations(
    entries: list[tuple[bytes, str]], target_lang: str | None
) -> tuple[list[tuple[bytes, str]], set[bytes]]:
    """Adds the UI strings and reports which keys may be substituted alone."""
    if not target_lang:
        return entries, set()

    language = "EN" if target_lang.upper().startswith("EN") else target_lang.upper()
    translations = NAME_TRANSLATIONS.get(language)
    if translations is None:
        raise ValueError(f"Unsupported name translation language: {target_lang}")

    merged: OrderedDict[bytes, str] = OrderedDict(entries)
    standalone: set[bytes] = set()
    for source, translation in RUNTIME_MENU_TRANSLATIONS[language].items():
        key = source.encode("cp932")
        merged.setdefault(key, translation)
        standalone.add(key)

    for source, translation in translations.items():
        key = source.encode("cp932")
        merged[key] = translation
        standalone.add(key)

    padding = " " + chr(0x3000)
    for plate in NAME_PLATES:
        translation = translations.get(plate.strip(padding))
        if translation is None:
            continue
        indent = plate[: len(plate) - len(plate.lstrip(padding))]
        key = plate.encode("cp932")
        merged[key] = indent + translation
        standalone.add(key)

    return list(merged.items()), standalone


def write_overlay(
    entries: list[tuple[bytes, str]], output: Path, standalone: set[bytes] | None = None
) -> dict:
    char_to_byte, byte_to_char = build_encoding()
    unmapped: Counter = Counter()

    codepage = bytearray(512)
    for code, character in byte_to_char.items():
        if character in MAPPED_CHARACTERS:
            struct.pack_into("<H", codepage, code * 2, ord(character))

    payload = bytearray(OVERLAY_MAGIC)
    payload.extend(struct.pack("<I", len(entries)))
    payload.extend(codepage)

    seen: set[bytes] = set()
    lengths: list[int] = []
    standalone = standalone or set()
    for source, translation in entries:
        if source in seen:
            raise ValueError("Duplicate source byte sequence in overlay")
        seen.add(source)
        encoded = encode_translation(translation, char_to_byte, unmapped)
        lengths.append(len(encoded))
        flags = STANDALONE_FLAG if source in standalone else 0
        payload.extend(struct.pack("<III", len(source), len(encoded), flags))
        payload.extend(source)
        payload.extend(encoded)

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(payload)

    lengths.sort()
    return {
        "entries": len(entries),
        "standalone": len(standalone),
        "codepage": len(MAPPED_CHARACTERS),
        "unmapped": unmapped,
        "longest": lengths[-1] if lengths else 0,
        "over_90": sum(1 for length in lengths if length > 90),
    }


@dataclass(frozen=True)
class FittedLine:
    source: str
    translation: str
    width: int
    standalone: bool


def measure_entries(
    entries: list[tuple[bytes, str]],
    standalone: set[bytes] | None = None,
    metrics: delta_fit.Metrics | None = None,
) -> list[FittedLine]:
    """Widths of every line as the proxy will draw it.

    The proxy decodes each stored byte through the overlay codepage and
    measures that character, so the text is round-tripped here rather than
    measured as written: a character with no code of its own reaches the screen
    as '?' and is that wide.
    """
    char_to_byte, byte_to_char = build_encoding()
    standalone = standalone or set()
    discard: Counter = Counter()

    fitted: list[FittedLine] = []
    with delta_fit.Fitter(metrics) as fitter:
        plate_step = fitter.metrics.name_plate_step
        for source, translation in entries:
            encoded = encode_translation(translation, char_to_byte, discard)
            drawn = "".join(byte_to_char.get(code, chr(code)) for code in encoded)
            is_plate = source in standalone
            fitted.append(
                FittedLine(
                    source=source.decode("cp932", errors="replace"),
                    translation=drawn,
                    width=fitter.width(drawn, plate_step if is_plate else 0),
                    standalone=is_plate,
                )
            )
    return fitted


def overflowing(lines: list[FittedLine], metrics: delta_fit.Metrics) -> list[FittedLine]:
    """Scenario lines that run past the right edge of the message frame.

    Standalone strings are left out. They are name plates and menu captions
    drawn outside this frame, on geometry this module does not know, so calling
    them too wide would be a guess dressed up as a measurement.
    """
    limit = metrics.available_width
    return sorted(
        (line for line in lines if not line.standalone and line.width > limit),
        key=lambda line: line.width,
        reverse=True,
    )


def print_fit(lines: list[FittedLine], metrics: delta_fit.Metrics) -> None:
    scenario = [line for line in lines if not line.standalone]
    if not scenario:
        return
    widest = max(line.width for line in scenario)
    over = overflowing(lines, metrics)
    print(
        f"Widest line: {widest} px in a {metrics.available_width} px frame "
        f"(text from x={metrics.text_x}, {metrics.face} at {metrics.font_height} px)"
    )
    if not over:
        print("Lines wider than the frame: none")
        return
    print(f"Lines wider than the frame: {len(over)}")
    for line in over[:5]:
        print(f"  {line.width:5d} px  {line.translation[:60]}")
    if len(over) > 5:
        print(f"  ... and {len(over) - 5} more; run 'fit' for the full list")


def write_fit_report(lines: list[FittedLine], metrics: delta_fit.Metrics, output: Path) -> int:
    """Writes every overflowing line as TSV. Returns how many there were."""
    over = overflowing(lines, metrics)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        handle.write("Width\tOverflow\tOriginal\tTranslation\n")
        for line in over:
            handle.write(
                f"{line.width}\t{line.width - metrics.available_width}\t"
                f"{line.source}\t{line.translation}\n"
            )
    return len(over)


def read_overlay(path: Path) -> list[tuple[bytes, str, bool]]:
    data = path.read_bytes()
    if len(data) < 8 + 512 or data[:4] != OVERLAY_MAGIC:
        raise ValueError(f"Not a {OVERLAY_MAGIC.decode()} overlay")

    count = struct.unpack_from("<I", data, 4)[0]
    position = 8
    codepage = {}
    for code in range(256):
        character = struct.unpack_from("<H", data, position + code * 2)[0]
        if character:
            codepage[code] = chr(character)
    position += 512

    entries: list[tuple[bytes, str, bool]] = []
    for _ in range(count):
        if position + 12 > len(data):
            raise ValueError("Truncated overlay header")
        source_size, translation_size, flags = struct.unpack_from("<III", data, position)
        position += 12
        if position + source_size + translation_size > len(data):
            raise ValueError("Truncated overlay entry")
        source = data[position : position + source_size]
        position += source_size
        encoded = data[position : position + translation_size]
        position += translation_size
        translation = "".join(
            codepage.get(byte, chr(byte)) for byte in encoded
        )
        entries.append((source, translation, bool(flags & STANDALONE_FLAG)))
    if position != len(data):
        raise ValueError("Trailing data in overlay")
    return entries


def command_extract(args: argparse.Namespace) -> None:
    digest = verify_source(args.source, args.allow_unknown)
    entries = extract_entries(args.source)
    write_workbook(entries, args.output, digest)
    print(f"Extracted unique Japanese strings: {len(entries)}")
    windows = {
        location.split()[0] for entry in entries for location in entry.dialog_lines
    }
    print(f"Dialog windows: {len(windows)}")
    # Worth its own line: an option missing from the workbook is an option about
    # to stay Japanese on screen, and no coverage figure would show it, because
    # coverage counts what the workbook holds.
    data = args.source.read_bytes()
    options = option_offsets(data)
    menus = sum(1 for block in find_choice_blocks(data) if block.branches)
    extracted = {offset for entry in entries for offset in entry.offsets}
    labelled = sum(1 for name in windows if name.startswith("C"))
    print(
        f"Choice menus: {menus} with {len(options)} options, "
        f"{len(options - extracted)} missing from the workbook"
    )
    print(f"Choice windows: {labelled} labelled C, menus and echoes of their lines")
    print(f"Japanese characters: {sum(len(entry.text) for entry in entries)}")
    print(f"Workbook: {args.output.resolve()}")


def command_build(args: argparse.Namespace) -> None:
    print_coverage(workbook_coverage(args.workbook))

    entries = workbook_translations(args.workbook)
    entries, standalone = add_name_translations(entries, args.target_lang)
    if not entries:
        raise ValueError("The TL column has no translations")

    # Measured before anything is written, so --strict-fit leaves no overlay
    # behind for the next step to pick up and install.
    metrics = delta_fit.read_metrics(args.launcher_ini)
    reason = delta_fit.unavailable_reason(metrics)
    lines = None if reason else measure_entries(entries, standalone, metrics)
    if lines is not None and args.strict_fit:
        over = overflowing(lines, metrics)
        if over:
            raise ValueError(
                f"{len(over)} lines do not fit the message frame; nothing was written"
            )

    report = write_overlay(entries, args.output, standalone)
    print(f"Overlay entries: {report['entries']}")
    print(f"Overlay codepage characters: {report['codepage']}")
    print(f"Standalone UI strings: {report['standalone']}")

    if lines is None:
        # Without GDI the only thing left is the half-width cell count, which
        # describes the Japanese grid rather than the drawn line. Say so
        # instead of presenting it as a width.
        print(f"Line widths not measured: {reason}")
        print(f"Longest encoded line: {report['longest']} half-width cells (estimate)")
    else:
        print_fit(lines, metrics)
        if args.fit_report is not None:
            count = write_fit_report(lines, metrics, args.fit_report)
            print(f"Fit report: {args.fit_report.resolve()} ({count} lines)")

    warn_unmapped(report["unmapped"])
    print(f"Overlay: {args.output.resolve()}")


def warn_unmapped(unmapped: Counter) -> None:
    if not unmapped:
        return
    preview = ", ".join(
        f"{character!r}x{count}" for character, count in unmapped.most_common(20)
    )
    print(
        f"WARNING: unsupported characters replaced with '?': {preview}",
        file=sys.stderr,
    )


def command_smoke(args: argparse.Namespace) -> None:
    verify_source(args.source, args.allow_unknown)
    entries = extract_entries(args.source)
    if not entries:
        raise ValueError("No Japanese strings found")
    try:
        entry = entries[args.entry_index]
    except IndexError as exc:
        raise ValueError(
            f"Smoke entry index {args.entry_index} is outside {len(entries)} entries"
        ) from exc
    write_overlay([(entry.source, args.translation)], args.output)
    print(f"Original: {entry.text}")
    print(f"Translation: {args.translation}")
    print(f"Overlay: {args.output.resolve()}")


def command_coverage(args: argparse.Namespace) -> None:
    print_coverage(workbook_coverage(args.workbook))


def command_fit(args: argparse.Namespace) -> None:
    """Measures an overlay that already exists, or a workbook before one does."""
    metrics = delta_fit.read_metrics(args.launcher_ini)
    reason = delta_fit.unavailable_reason(metrics)
    if reason is not None:
        raise ValueError(reason)

    if args.input.suffix.casefold() == ".xlsx":
        entries = workbook_translations(args.input)
        entries, standalone = add_name_translations(entries, args.target_lang)
    else:
        stored = read_overlay(args.input)
        entries = [(source, translation) for source, translation, _ in stored]
        standalone = {source for source, _, flag in stored if flag}

    lines = measure_entries(entries, standalone, metrics)
    print_fit(lines, metrics)
    if args.report is not None:
        count = write_fit_report(lines, metrics, args.report)
        print(f"Fit report: {args.report.resolve()} ({count} lines)")


def command_validate(args: argparse.Namespace) -> None:
    entries = read_overlay(args.overlay)
    print(f"Valid {OVERLAY_MAGIC.decode()} overlay entries: {len(entries)}")
    if entries:
        print(f"First source: {entries[0][0].decode('cp932', errors='replace')}")
        print(f"First translation: {entries[0][1]}")
        print(f"Standalone entries: {sum(1 for entry in entries if entry[2])}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract Delta RSAN.SD text and build a runtime translation overlay."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    extract = subparsers.add_parser("extract")
    extract.add_argument("source", type=Path)
    extract.add_argument("output", type=Path)
    extract.add_argument("--allow-unknown", action="store_true")
    extract.set_defaults(func=command_extract)

    build = subparsers.add_parser("build-overlay")
    build.add_argument("workbook", type=Path)
    build.add_argument("output", type=Path)
    build.add_argument("--target-lang", choices=sorted(NAME_TRANSLATIONS))
    build.add_argument(
        "--launcher-ini",
        type=Path,
        help="delta_launcher.ini to read the layout from; defaults to the proxy's own defaults.",
    )
    build.add_argument(
        "--fit-report", type=Path, help="Write every overflowing line to this TSV."
    )
    build.add_argument(
        "--strict-fit",
        action="store_true",
        help="Fail instead of warning when a line does not fit the message frame.",
    )
    build.set_defaults(func=command_build)

    coverage = subparsers.add_parser(
        "coverage", help="Report how much of a workbook is translated."
    )
    coverage.add_argument("workbook", type=Path)
    coverage.set_defaults(func=command_coverage)

    fit = subparsers.add_parser(
        "fit", help="Measure translated lines against the message window."
    )
    fit.add_argument("input", type=Path, help="A .xlsx workbook or a built overlay.")
    fit.add_argument("--target-lang", choices=sorted(NAME_TRANSLATIONS))
    fit.add_argument("--launcher-ini", type=Path)
    fit.add_argument("--report", type=Path, help="Write every overflowing line to this TSV.")
    fit.set_defaults(func=command_fit)

    smoke = subparsers.add_parser("smoke-overlay")
    smoke.add_argument("source", type=Path)
    smoke.add_argument("output", type=Path)
    smoke.add_argument(
        "--translation",
        default="Из окна вагона видны голые зимние деревья.",
    )
    smoke.add_argument("--entry-index", type=int, default=0)
    smoke.add_argument("--allow-unknown", action="store_true")
    smoke.set_defaults(func=command_smoke)

    validate = subparsers.add_parser("validate-overlay")
    validate.add_argument("overlay", type=Path)
    validate.set_defaults(func=command_validate)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        args.func(args)
    except (OSError, ValueError, KeyError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
