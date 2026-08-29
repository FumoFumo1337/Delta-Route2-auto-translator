"""Proofreading a choice menu, and reading its Dialog Line downstream.

A rule has to reach a C window exactly as it reaches a message window: every
`dialogs` rule addresses a window by its letter, and the letter changed.

The tools that parse Dialog Line have to accept it as well. delta_deepl refuses
a row whose reference it cannot read, so a C window it did not recognise would
have stopped translation outright rather than mistranslating anything.
"""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from context import requires_openpyxl  # noqa: F401  (prepares sys.path)

import delta_deepl
import proofread


@requires_openpyxl
class TestProofreadingAMenu(unittest.TestCase):
    """A rule has to reach a menu exactly as it reaches a message window.

    The letter changed, and every `dialogs` rule addresses a window by it.
    """

    def workbook(self, rows: list[tuple[str, str, str]]) -> Path:
        import openpyxl

        directory = Path(self.enterContext(tempfile.TemporaryDirectory()))
        path = directory / "book.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Scenario"
        sheet.append(["ID", "Dialog Line", "Original", "TL"])
        for index, row in enumerate(rows, start=1):
            sheet.append([index, *row])
        book.save(path)
        return path

    def run_proofread(
        self, rows: list[tuple[str, str, str]], rules: dict
    ) -> list[str]:
        import openpyxl

        book = self.workbook(rows)
        rules_path = book.with_name("rules.json")
        rules_path.write_text(json.dumps(rules, ensure_ascii=False), encoding="utf-8")
        output = book.with_name("out.xlsx")
        proofread.proofread(book, output, rules_path, None)
        sheet = openpyxl.load_workbook(output)["Scenario"]
        return [sheet.cell(row, 4).value for row in range(2, sheet.max_row + 1)]

    OPTIONS = [
        ("C00011 1/2", "Ｄｏｌｌとしてはまだまだ…", "как для Куклы она пока ещё… женщина, с"),
        ("C00011 2/2", "虐め甲斐のある女だ", "которой стоит поиздеваться"),
    ]

    def test_a_rule_keyed_by_a_choice_id_rewrites_both_options(self) -> None:
        """The wording being repaired here is the defect itself: one sentence
        cut across the two alternatives instead of two standalone options."""
        self.assertEqual(
            self.run_proofread(
                self.OPTIONS,
                {
                    "dialogs": {
                        "C00011": [
                            "Как Кукла она ещё далека от идеала",
                            "Такую женщину стоит помучить",
                        ]
                    }
                },
            ),
            ["Как Кукла она ещё далека от идеала", "Такую женщину стоит помучить"],
        )

    def test_a_signature_rule_finds_a_menu_by_its_japanese(self) -> None:
        """Signature rules are why relabelling a window did not invalidate any
        existing rule: the number is only a label."""
        self.assertEqual(
            self.run_proofread(
                self.OPTIONS,
                {
                    "dialogs": {
                        "D02211": {
                            "original": [row[1] for row in self.OPTIONS],
                            "translation": ["Первый вариант", "Второй вариант"],
                        }
                    }
                },
            ),
            ["Первый вариант", "Второй вариант"],
        )

    def test_a_signature_rule_keeps_working_when_the_choice_is_renumbered(self) -> None:
        """The C number follows plot order and is not the rule identity."""
        self.assertEqual(
            self.run_proofread(
                [
                    ("C00421 1/2", self.OPTIONS[0][1], "черновик 1"),
                    ("C00421 2/2", self.OPTIONS[1][1], "черновик 2"),
                ],
                {
                    "dialogs": {
                        "D02211": {
                            "original": [row[1] for row in self.OPTIONS],
                            "translation": ["Вариант один", "Вариант два"],
                        }
                    }
                },
            ),
            ["Вариант один", "Вариант два"],
        )

    def test_a_menu_and_a_message_colliding_on_one_row_name_both(self) -> None:
        """Two of the 29 option rows are also a line of a message window."""
        with self.assertRaises(ValueError) as caught:
            self.run_proofread(
                [("C00002 1/2; D00579 1/1", "麗佳の苦痛に彩られた声…", "черновик")],
                {
                    "dialogs": {
                        "C00002": ["как вариант"],
                        "D00579": ["как реплика"],
                    }
                },
            )
        message = str(caught.exception)
        self.assertIn("C00002 1", message)
        self.assertIn("D00579 1", message)


class TestDownstreamReferences(unittest.TestCase):
    """The tools that read Dialog Line have to accept the new letter.

    delta_deepl refuses a row whose Dialog Line it cannot parse, so a C window
    would have stopped translation outright.
    """

    def test_the_translator_parses_a_choice_reference(self) -> None:
        self.assertEqual(
            delta_deepl.parse_dialog_refs("C00001 1/2"), [("C00001", 1, 2)]
        )

    def test_the_translator_parses_a_row_in_a_menu_and_a_message(self) -> None:
        self.assertEqual(
            delta_deepl.parse_dialog_refs("C00002 1/2; D00579 1/1"),
            [("C00002", 1, 2), ("D00579", 1, 1)],
        )

    def test_proofread_parses_a_choice_reference(self) -> None:
        self.assertEqual(
            proofread.DIALOG_REF_RE.findall("C00010 2/2; D01863 1/1"),
            [("C00010", "2", "2"), ("D01863", "1", "1")],
        )
