"""Layering and application order of the proofread rule files.

Two layers are loaded per target language: the reusable technical profile under
profiles/, then the per-game file under work/. The project layer is applied
second and therefore wins, which is what makes it safe to keep character names
out of the shared profile.
"""

from __future__ import annotations

import json
import io
import tempfile
import unittest
from contextlib import redirect_stderr
from pathlib import Path

from context import TOOLS_ROOT, requires_openpyxl  # noqa: F401  (prepares sys.path)

import proofread


class TestMergeRules(unittest.TestCase):
    def test_project_layer_overrides_a_scalar(self) -> None:
        merged = proofread.merge_rules({"a": 1}, {"a": 2})
        self.assertEqual(merged["a"], 2)

    def test_maps_merge_recursively(self) -> None:
        merged = proofread.merge_rules(
            {"replacements": {"one": "1", "two": "2"}},
            {"replacements": {"two": "II", "three": "3"}},
        )
        self.assertEqual(
            merged["replacements"], {"one": "1", "two": "II", "three": "3"}
        )

    def test_lists_append_in_layer_order(self) -> None:
        """Regex order matters, and the project's rules run after the profile's."""
        merged = proofread.merge_rules(
            {"regex_replacements": [{"pattern": "a"}]},
            {"regex_replacements": [{"pattern": "b"}]},
        )
        self.assertEqual(
            [item["pattern"] for item in merged["regex_replacements"]], ["a", "b"]
        )

    def test_base_is_not_mutated(self) -> None:
        base = {"replacements": {"one": "1"}}
        proofread.merge_rules(base, {"replacements": {"one": "changed"}})
        self.assertEqual(base["replacements"]["one"], "1")


class TestApplyRuleBlock(unittest.TestCase):
    def test_plain_replacements(self) -> None:
        result = proofread.apply_rule_block(
            "源", "Хеллоу мир", {"replacements": {"Хеллоу": "Привет"}}
        )
        self.assertEqual(result, "Привет мир")

    def test_regex_runs_after_plain_replacements(self) -> None:
        rules = {
            "replacements": {"aaa": "bbb"},
            "regex_replacements": [{"pattern": "b+", "replacement": "B"}],
        }
        self.assertEqual(proofread.apply_rule_block("s", "aaa", rules), "B")

    def test_regex_order_is_preserved(self) -> None:
        rules = {
            "regex_replacements": [
                {"pattern": "oo", "replacement": "u"},
                {"pattern": "u+", "replacement": "U"},
            ]
        }
        self.assertEqual(proofread.apply_rule_block("s", "oo", rules), "U")

    def test_ignore_case_flag(self) -> None:
        rules = {
            "regex_replacements": [
                {"pattern": "abc", "replacement": "x", "ignore_case": True}
            ]
        }
        self.assertEqual(proofread.apply_rule_block("s", "ABC", rules), "x")

    def test_source_rule_fires_only_on_a_matching_source(self) -> None:
        rules = {
            "source_rules": [
                {"contains": "麗佳", "replacements": {"она": "Рейка"}}
            ]
        }
        self.assertEqual(
            proofread.apply_rule_block("「麗佳！」", "она здесь", rules), "Рейка здесь"
        )
        self.assertEqual(
            proofread.apply_rule_block("「玲！」", "она здесь", rules), "она здесь"
        )

    def test_malformed_regex_rule_is_rejected(self) -> None:
        """A typo must fail loudly rather than quietly change nothing."""
        with self.assertRaises(ValueError):
            proofread.apply_rule_block("s", "text", {"regex_replacements": [{"x": 1}]})

    def test_empty_rules_leave_the_cell_alone(self) -> None:
        self.assertEqual(proofread.apply_rule_block("s", "текст", {}), "текст")

    def test_ellipsis_repetition_preserves_case(self) -> None:
        profile = proofread.load_rules(
            TOOLS_ROOT / "profiles" / "proofread_common.ru.json"
        )
        self.assertEqual(
            proofread.apply_rule_block(
                "s", "ммм… ммм… ммм… ммм… ммм… ммм…", profile
            ),
            "ммм… ммм… ммм…",
        )
        mixed_case = "Я… я… я… я… я…"
        self.assertEqual(
            proofread.apply_rule_block("s", mixed_case, profile), "Я… я… я… я…"
        )


@requires_openpyxl
class TestDialogRuleConflicts(unittest.TestCase):
    """Identical lines are deduplicated into one workbook row, so a row can
    belong to several windows. Two `dialogs` rules that reach the same row want
    different bytes in one place, which no rewrite of that row can satisfy."""

    def workbook(self, dialog_line: str) -> Path:
        import openpyxl

        directory = Path(self.enterContext(tempfile.TemporaryDirectory()))
        path = directory / "book.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Scenario"
        sheet.append(["ID", "Dialog Line", "Original", "TL"])
        sheet.append([1, dialog_line, "宙吊りを見ている", "черновик"])
        book.save(path)
        return path

    def rules(self, dialogs: dict) -> Path:
        directory = Path(self.enterContext(tempfile.TemporaryDirectory()))
        path = directory / "rules.json"
        path.write_text(json.dumps({"dialogs": dialogs}), encoding="utf-8")
        return path

    def run_proofread(self, dialog_line: str, dialogs: dict) -> Path:
        book = self.workbook(dialog_line)
        output = book.with_name("out.xlsx")
        proofread.proofread(book, output, self.rules(dialogs), None)
        return output

    def cell(self, path: Path) -> str:
        import openpyxl

        return openpyxl.load_workbook(path)["Scenario"].cell(2, 4).value

    def test_one_rule_reaching_a_shared_row_is_applied(self) -> None:
        output = self.run_proofread("D1 1/2; D2 1/2", {"D1": ["Видела подвешивание"]})
        self.assertEqual(self.cell(output), "Видела подвешивание")

    def test_agreeing_rules_are_not_a_conflict(self) -> None:
        """Two windows may legitimately want the same wording."""
        output = self.run_proofread(
            "D1 1/2; D2 1/2", {"D1": ["одно и то же"], "D2": ["одно и то же"]}
        )
        self.assertEqual(self.cell(output), "одно и то же")

    def test_disagreeing_rules_name_both_sides(self) -> None:
        """The row number alone leaves the reader hunting through the rules for
        which two of them collided."""
        with self.assertRaises(ValueError) as caught:
            self.run_proofread(
                "D1 1/2; D2 1/2", {"D1": ["Видела Рей"], "D2": ["Видела Джона"]}
            )
        message = str(caught.exception)
        self.assertIn("D1 1", message)
        self.assertIn("D2 1", message)
        self.assertIn("Видела Рей", message)
        self.assertIn("Видела Джона", message)

    def test_positions_are_matched_within_the_window(self) -> None:
        """A rule for line 2 must not fire on line 1 of the same window."""
        output = self.run_proofread("D1 1/2", {"D1": ["первая", "вторая"]})
        self.assertEqual(self.cell(output), "первая")

    def test_source_signature_survives_dialog_renumbering(self) -> None:
        import openpyxl

        directory = Path(self.enterContext(tempfile.TemporaryDirectory()))
        book_path = directory / "book.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Scenario"
        sheet.append(["ID", "Dialog Line", "Original", "TL"])
        sheet.append([1, "D99 1/2", "一行目", "draft one"])
        sheet.append([2, "D99 2/2", "二行目", "draft two"])
        book.save(book_path)

        rules = self.rules(
            {
                "D12": {
                    "original": ["一行目", "二行目"],
                    "translation": ["первая", "вторая"],
                }
            }
        )
        output = directory / "out.xlsx"
        proofread.proofread(book_path, output, rules, None)
        result = openpyxl.load_workbook(output)["Scenario"]
        self.assertEqual(result.cell(2, 4).value, "первая")
        self.assertEqual(result.cell(3, 4).value, "вторая")

    def test_stale_signature_fails_instead_of_touching_another_window(self) -> None:
        with self.assertRaisesRegex(ValueError, "no matching source window"):
            self.run_proofread(
                "D1 1/1",
                {
                    "D1": {
                        "original": ["другой оригинал"],
                        "translation": ["не применять"],
                    }
                },
            )

    def test_failed_validation_does_not_leave_a_backup(self) -> None:
        book = self.workbook("D1 1/1")
        backup = book.with_name("out.xlsx.before.xlsx")
        rules = self.rules(
            {
                "D1": {
                    "original": ["different original"],
                    "translation": ["do not apply"],
                }
            }
        )
        with self.assertRaisesRegex(ValueError, "no matching source window"):
            proofread.proofread(book, book.with_name("out.xlsx"), rules, backup)
        self.assertFalse(backup.exists())

    def test_expanded_window_is_named_in_stale_signature_error(self) -> None:
        import openpyxl

        directory = Path(self.enterContext(tempfile.TemporaryDirectory()))
        book_path = directory / "book.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Scenario"
        sheet.append(["ID", "Dialog Line", "Original", "TL"])
        sheet.append([1, "D99 1/2", "new first line", "draft one"])
        sheet.append([2, "D99 2/2", "old line", "draft two"])
        book.save(book_path)
        rules_path = self.rules(
            {
                "D12": {
                    "original": ["old line"],
                    "translation": ["reviewed line"],
                }
            }
        )

        with self.assertRaises(ValueError) as caught:
            proofread.proofread(book_path, directory / "out.xlsx", rules_path, None)
        message = str(caught.exception)
        self.assertIn("D99", message)
        self.assertIn("new first line", message)
        self.assertIn("Update both original and translation arrays", message)


class TestRuleKinds(unittest.TestCase):
    def test_no_speaker_keyed_rule_kind_exists(self) -> None:
        """A source line is unique across the workbook, so `sources` already
        addresses one row exactly. An experimental speaker-keyed kind was
        removed; this keeps it from drifting back in unnoticed."""
        source = (TOOLS_ROOT / "py" / "proofread.py").read_text(encoding="utf-8")
        self.assertNotIn("speaker_rules", source)


@requires_openpyxl
class TestProofreadLoopWarning(unittest.TestCase):
    def workbook(self, translation: str) -> Path:
        import openpyxl

        directory = Path(self.enterContext(tempfile.TemporaryDirectory()))
        path = directory / "book.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Scenario"
        sheet.append(["ID", "Dialog Line", "Original", "TL"])
        sheet.append([1, "D00001 1/1", "止めて", translation])
        book.save(path)
        return path

    def test_remaining_loop_is_a_cli_warning(self) -> None:
        source = self.workbook("Stop it, " * 80)
        errors = io.StringIO()
        with redirect_stderr(errors):
            proofread.proofread(source, source.with_name("out.xlsx"), None, None)
        warning = errors.getvalue()
        self.assertIn("WARNING: unresolved looped translations", warning)
        self.assertIn("D00001", warning)

    def test_normal_dialog_has_no_warning(self) -> None:
        source = self.workbook("Stop it, please.")
        errors = io.StringIO()
        with redirect_stderr(errors):
            proofread.proofread(source, source.with_name("out.xlsx"), None, None)
        self.assertEqual(errors.getvalue(), "")


if __name__ == "__main__":
    unittest.main()
