"""Small workbook fixture spanning extraction, coverage, and proofreading.

Unlike the focused tests that build a one-row workbook in memory, this pins the
actual nine-column interchange format. It deliberately combines deduplication,
branching windows, a renumbered dialog signature, partial coverage, and the
machine-translation repetition edge cases in one resource.
"""

from __future__ import annotations

import base64
import json
import tempfile
import unittest
from pathlib import Path

from context import FIXTURES_DIR, TOOLS_ROOT, requires_openpyxl

import delta_overlay as overlay
import proofread


FIXTURE = FIXTURES_DIR / "scenario_edge_cases.xlsx"


@requires_openpyxl
class TestScenarioWorkbookFixture(unittest.TestCase):
    def scenario_rows(self) -> list[dict[str, object]]:
        import openpyxl

        workbook = openpyxl.load_workbook(FIXTURE, read_only=True)
        sheet = workbook["Scenario"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        rows = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        workbook.close()
        return rows

    def test_fixture_pins_the_extractor_workbook_contract(self) -> None:
        rows = self.scenario_rows()
        self.assertEqual([row["ID"] for row in rows], list(range(1, 12)))
        self.assertEqual(len({row["Original"] for row in rows}), len(rows))

        shared = rows[0]
        self.assertEqual(shared["Occurrences"], 2)
        self.assertEqual(
            shared["Dialog Line"], "D00100 1/2; D00200 1/2"
        )
        for row in rows:
            source = base64.b64decode(str(row["SourceBytesBase64"]), validate=True)
            self.assertEqual(source.decode("cp932"), row["Original"])

    def test_extractor_filter_cases_are_cp932_bytes(self) -> None:
        import openpyxl

        workbook = openpyxl.load_workbook(FIXTURE, read_only=True)
        sheet = workbook["ExtractorCases"]
        for case, encoded, decoded, expected, _reason in sheet.iter_rows(
            min_row=2, values_only=True
        ):
            with self.subTest(case=case):
                source = base64.b64decode(encoded, validate=True)
                self.assertEqual(source.decode("cp932"), decoded)
                self.assertEqual(overlay.looks_like_scenario_text(decoded), expected)
        workbook.close()

    def test_coverage_counts_shared_and_partial_windows(self) -> None:
        report = overlay.workbook_coverage(FIXTURE)
        self.assertEqual((report.rows, report.translated), (11, 9))
        self.assertEqual(report.windows, 8)
        self.assertEqual(
            (report.complete_windows, report.partial_windows, report.empty_windows),
            (6, 1, 1),
        )

    def test_proofread_keeps_distinct_branches_and_rebases_a_signature(self) -> None:
        import openpyxl

        directory = Path(self.enterContext(tempfile.TemporaryDirectory()))
        rules_path = directory / "rules.json"
        output_path = directory / "proofread.xlsx"
        rules = {
            "dialogs": {
                "OLD_BRANCH_A": {
                    "original": ["共通の選択肢", "枝分かれA"],
                    "translation": ["Общий выбор", "Ветка A"],
                },
                "OLD_BRANCH_B": {
                    "original": ["共通の選択肢", "枝分かれB"],
                    "translation": ["Общий выбор", "Ветка B"],
                },
                "D00001": {
                    "original": ["番号が変わる一行目", "番号が変わる二行目"],
                    "translation": ["Перенумерованная первая", "Перенумерованная вторая"],
                },
            }
        }
        rules_path.write_text(json.dumps(rules, ensure_ascii=False), encoding="utf-8")
        proofread.proofread(
            FIXTURE,
            output_path,
            rules_path,
            None,
            TOOLS_ROOT / "profiles" / "proofread_common.ru.json",
        )

        workbook = openpyxl.load_workbook(output_path, read_only=True)
        sheet = workbook["Scenario"]
        translations = {
            str(row[4]): str(row[5] or "")
            for row in sheet.iter_rows(min_row=2, values_only=True)
        }
        workbook.close()
        self.assertEqual(translations["共通の選択肢"], "Общий выбор")
        self.assertEqual(translations["枝分かれA"], "Ветка A")
        self.assertEqual(translations["枝分かれB"], "Ветка B")
        self.assertEqual(
            translations["番号が変わる一行目"], "Перенумерованная первая"
        )
        self.assertEqual(
            translations["番号が変わる二行目"], "Перенумерованная вторая"
        )
        self.assertEqual(translations["長い息遣い"], "ммм… ммм… ммм…")
        self.assertEqual(translations["大文字と小文字"], "Я… я… я… я…")
        self.assertIn("настоящая длинная фраза", translations["長い本物の文章"])


if __name__ == "__main__":
    unittest.main()
